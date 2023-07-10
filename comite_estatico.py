import sys
import warnings
import seaborn as sns
import matplotlib.pyplot as plt
import collections

from sklearn.preprocessing import StandardScaler #importar preprocessamento para normalizar dados,
from sklearn.calibration import CalibratedClassifierCV
# Importing dataset and preprocessing routines
from sklearn.datasets import fetch_openml
from sklearn.ensemble import VotingClassifier

from sklearn.linear_model import Perceptron
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier

from deslib.des import KNORAU

from deslib.static import StackedClassifier
from sklearn.model_selection import GridSearchCV

print("Comite estatico em NOVA FUNCAO")


if not sys.warnoptions:
    warnings.simplefilter("ignore")

divergencia_classificadores = False

from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter


#linha_lida =  int(input('Por favor digite a linha: '))
linha_lida =  232
#linha_lida =  862
#coluna_valor = 8
#coluna_desvio_padrao = 9

coluna_valor = 2
coluna_desvio_padrao = 3

wb = load_workbook('comite estatico NOVA FUNCAO.xlsx')

ws = wb['estatico']


print("Comite Estatico")
#!/usr/bin/env python
# coding: utf-8
k_dynamic_combination = 0
from numpy import mean
from numpy import std
import numpy as np
import pandas  as pd
from sklearn.datasets import make_classification
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import RepeatedStratifiedKFold
from deslib.des.knora_e import KNORAE
from deslib.des.knora_u import KNORAU
from deslib.des.base import BaseDES

from deslib.util.aggregation import aggregate_proba_ensemble_weighted
from deslib.util.aggregation import get_weighted_votes

import deslib
deslib.__version__

from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter

#print()


# 3 Exemplos
# 2 CLASSES
# 5 Classificadores
'''
probabilities
[[[0. 1.] c1
  [1. 0.] c 2
  [1. 0.]
  [1. 0.]
  [0. 1.]]

ex 2
 [[0. 1.]
  [0. 1.]
  [1. 0.]
  [0. 1.]
  [0. 1.]]

 [[1. 0.]
  [0. 1.]
  [0. 1.]
  [1. 0.]
  [0. 1.]]]
'''

'''
Aqui
[[0.95516666 0.04483334]
 [0.98281892 0.01718108]
 [0.99720851 0.00279149]
 [0.0515     0.9485    ]
 [0.02384775 0.97615225]
 [0.00945816 0.99054184]]
'''
'''
VOTES
[[1 -- 0 -- 1]
 [-- 1 0 0 1]
 [0 1 1 0 --]]
'''


#porcentagem_escolheu_MAX, porcentagem_escolheu_SOFT, porcentagem_escolheu_HARD, porcentagem_escolheu_MIN, porcentagem_escolheu_G_MEAN, porcentagem_escolheu_sum_weight , porcentagem_escolheu_rede_neural, porcentagem_escolheu_rede_neural_soft , porcentagem_escolheu_rede_neural_soft_div , porcentagem_escolheu_borda , porcentagem_escolheu_naive_bayes , porcentagem_escolheu_peso_ponderado_comite = [],[],[],[], [],[],[],[],[], [],[],[]
#porcentagem_escolheu_MAX= porcentagem_escolheu_SOFT= porcentagem_escolheu_HARD= porcentagem_escolheu_MIN= porcentagem_escolheu_G_MEAN= porcentagem_escolheu_sum_weight = porcentagem_escolheu_rede_neural= porcentagem_escolheu_rede_neural_soft = porcentagem_escolheu_rede_neural_soft_div = porcentagem_escolheu_borda = porcentagem_escolheu_naive_bayes = porcentagem_escolheu_peso_ponderado_comite = [[],[],[]]
porcentagem_escolheu_MAX= [[],[],[]]
porcentagem_escolheu_SOFT= [[],[],[]]
porcentagem_escolheu_HARD= [[],[],[]]
porcentagem_escolheu_MIN= [[],[],[]]
porcentagem_escolheu_G_MEAN= [[],[],[]]
porcentagem_escolheu_sum_weight = [[],[],[]]
porcentagem_escolheu_rede_neural= [[],[],[]]
porcentagem_escolheu_rede_neural_soft = [[],[],[]]
porcentagem_escolheu_rede_neural_soft_div = [[],[],[]]
porcentagem_escolheu_borda = [[],[],[]]
porcentagem_escolheu_naive_bayes = [[],[],[]]
porcentagem_escolheu_peso_ponderado_comite = [[],[],[]]
porcentagem_escolha_compartilhada = [[],[],[]]

geral_empate_classificadores_no_inicio = [[],[],[]]
geral_empate_classificadores_no_final = [[],[],[]]
resultados_hard = []
resultados_soft = []
resultados_max = []
resultados_min = []
resultados_geometric_mean = []
quantidade_exemplos_divergencia = []
quantidade_classificadores_selecionados = []
pesos_dos_classificadores_selecionados = []
#resultados_peso_ponderado_classe_cada_amostra_sem_ajustes = []
#resultados_peso_ponderado_classe_cada_amostra_ajustado_0a1 = []
#resultados_peso_ponderado_comite_classe_distancia_maxima_teste = []
resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1 = []
#resultados_sum_weight_votes_per_class = []
resultados_sum_weight_0a1_votes_per_class = []
#resultados_sum_weight_line_votes_per_class = []
resultados_sum_weight_0a1_line_votes_per_class = []

#resultados_dynamic_metric_fusionk1 = []
resultados_dynamic_metric_fusionk3= [[],[],[]]

#resultados_sum_weight  = []
#resultados_sum_weight_line  = []
resultados_escolheu_rede_neural  = []
resultados_escolheu_rede_neural_soft  = []
resultados_escolheu_rede_neural_soft_div  = []
resultados_escolheu_borda  = []
resultados_escolheu_naive_bayes  = []

resultados_maximo_na_combinacao = []

int_k_inserir = 0
rede_grid = True
rede_grid_estimador = None
rede_soft_grid_estimador = None


modelMAX = None
modelSOFT = None
modelHARD = None
modelMIN = None
modelG_MEAN = None
model_Rede_neural = None
modelRede_neural_soft = None
modelBorda = None
modelNaive_bayes = None
modelPeso_ponderado_comite = None
modelSum_weight = None
classificador_Rede_neural = None
classificador_neural_soft  = None
classificador_neural_softAUX  = None

knn_best_clf_in_grid = None
knn_is_grid = True


from sklearn.ensemble import VotingClassifier
from sklearn.utils.validation import check_is_fitted

class EnsembleClassifier(VotingClassifier):
    def __init__(
        self,
        estimators,
        *,
        voting="hard",
        weights=None,
        n_jobs=None,
        flatten_transform=True,
        verbose=False,
        voting_type='all',
    ):
        super(EnsembleClassifier, self).__init__(estimators=estimators)
        self.voting = voting
        self.weights = weights
        self.n_jobs = n_jobs
        self.flatten_transform = flatten_transform
        self.verbose = verbose

        self.voting_type = voting_type

        self.previsoes_classificadores = None
        self.probabilidades_classificadores = None
        self.previsoes_real = None

    def predict(self, X):
        """Predict class labels for X.
        Parameters
        ----------
        X : {array-like, sparse matrix} of shape (n_samples, n_features)
            The input samples.
        Returns
        -------
        maj : array-like of shape (n_samples,)
            Predicted class labels.
        """
        check_is_fitted(self)
        if self.voting == "soft":
            if self.voting_type == "rede_neural_soft" or self.voting_type == "rede_neural" or self.voting_type == 'naive_bayes' or self.voting_type == 'sum_weight_0a1_votes_per_class':
                self.previsoes_classificadores = self._predict(X)#Salvar alguns dados para usar na rede ou naive
            #print(" Chamou a função")
            proba = self.predict_proba(X)
            #print("proba")
            #print(proba)
            maj = np.argmax(proba, axis=1)
            #print("maj")
            #print(maj)

        else:  # 'HARD' voting
            predictions = self._predict(X)
            #print("predictions")
            #print(predictions)
            #print("No HARD")
            self.previsoes_classificadores = predictions

            maj = np.apply_along_axis(
                lambda x: np.argmax(np.bincount(x, weights=self._weights_not_none)),
                axis=1,
                arr=predictions,
            )

        maj = self.le_.inverse_transform(maj)#Transforma as classes clamadas de 0 ou 1 para seus nomes reais tipo true false

        return maj
    #



    def predict_proba(self, X):
        """Compute probabilities of possible outcomes for samples in X.
        Parameters
        ----------
        X : {array-like, sparse matrix} of shape (n_samples, n_features)
            The input samples.
        Returns
        -------
        avg : array-like of shape (n_samples, n_classes)
            Weighted average probability for each class per sample.
        """
        check_is_fitted(self)
        colecao_proba = self._collect_probas(X)
        #print("colecao_proba")
        #print(colecao_proba)

        if self.voting_type == 'soft':#SE SOFT (ou se for o HARD com voting_type='soft' usando apenas o predict_proba (ver as probabilidades) )
            avg = np.average(colecao_proba, axis=0, weights=self._weights_not_none)
            #print("O SOFT")

        else:#Se nao for HARD nem SOFT

            x, y, z = colecao_proba.shape
            #print(colecao_proba.shape)
            #print("############ PREPARAR OS DADOS PROS OUTROS ########################")
            
            #FORMATAR OS DADOS PARA USAR AS FUNÇÕES CRAIDAS
            novo_formato = np.empty([y, x, z])
            for k in range(y):#1 2 3 4 5 6
                #item = []
                for i in range(x):# 1 2 3
                    for l in range(z):# 1 2
                        #print(colecao_proba[i,k])
                        novo_formato[k,i,l] = colecao_proba[i,k,l]
                #novo_formato.append(item)

            self.probabilidades_classificadores = novo_formato
            self.previsoes_classificadores = self._predict(X)
            '''
            x, y, z = colecao_proba.shape
            print(colecao_proba.shape)
            print("####################################")
            novo_formato = []
            for k in range(y):#1 2 3 4 5 6
                item = []
                for i in range(x):# 1 2 3
                    print(colecao_proba[i,k])
                    item.append(colecao_proba[i,k])
                novo_formato.append(item)
            '''
            #print("####################################")
            #print("novo_formato")
            #print(novo_formato)

            #Escolha da técnica
            if self.voting_type == 'max':
                novo_formato = _max_proba(novo_formato)
                #print("novo_formato")
                #print(novo_formato)
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)

            elif self.voting_type == 'min':
                novo_formato = _minimun_proba(novo_formato)
                #print("novo_formato")
                #print(novo_formato)
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'rede_neural_soft':
                novo_formato = self.rede_neural_soft(novo_formato)
                #print("novo_formato")
                #print(novo_formato)
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'rede_neural':
                novo_formato = self.rede_neural_class(novo_formato)
                #print("novo_formato")
                #print(novo_formato)
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'geometric_mean':
                novo_formato = geometric_mean(novo_formato)
                #print("novo_formato")
                #print(novo_formato)
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1':
                novo_formato = self._peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1(novo_formato)
                #print("novo_formato")  _peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1
                #print(novo_formato)
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'sum_weight_0a1_votes_per_class':
                novo_formato = self._sum_weight_0a1_votes_per_class(self.previsoes_classificadores)
                #          votes = np.ma.MaskedArray(predictions, ~selected_classifiers)
                #          predicted_proba = self._sum_weight_0a1_votes_per_class(votes, self.n_classes_)
                #print("novo_formato")
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'borda':
                novo_formato = borda_class(novo_formato)
                #print("novo_formato") borda_class
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'naive_bayes':
                novo_formato = self.naive_bayes_combination(self.previsoes_classificadores, len(np.unique(y, return_counts=False)))
                #           votes = np.ma.MaskedArray(predictions, ~selected_classifiers)
                #           predicted_proba = self.naive_bayes_combination(votes, self.n_classes_)
                #print("novo_formato") naive_bayes_combination
                #print("novo_formato")
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)
            elif self.voting_type == 'dynamic_fusion':
                novo_formato = self.dynamic_metric_fusion2(novo_formato,v_total_k)
                #print("novo_formato")
                #print(novo_formato)
                avg = novo_formato# np.average(novo_formato, axis=0, weights=self._weights_not_none)

        return avg


    def dynamic_metric_fusion2(self, probabilities, vizinhos_val):#Devolve as probabilidades mas remove dos classificadores não competentes
        #print("----------Combinação dinamica 2---------")
        global divergencia_classificadores
        global knn_best_clf_in_grid
        divergencia_classificadores = True

        #print("VAlor k inserir é: " + str(int_k_inserir))

        #print("VAlor vizinhos_val é: " + str(vizinhos_val))




        #print("vizinhos_val")
        #print(vizinhos_val)
        #print("Meu X")
        #print(X_train)
        #print(len(X))
        total_exemplos = len(probabilities)
        #print("FUNÇÃO dynamic_metric_fusion")
        #print("probabilities")
        #print(probabilities)
        #print("selected_classifiers")
        #print(selected_classifiers)

        '''
        qnt_class_sel = 0
        linha,coluna = selected_classifiers.shape
        for lin in range(linha):
            for col in range(coluna):
                if selected_classifiers[lin,col] == True:
                    qnt_class_sel = qnt_class_sel + 1

        #qnt_class_sel = collections.Counter(selected_classifiers)[True]
        qnt_class_sel = qnt_class_sel / linha
        #print("Qnt sele clas")
        #print(qnt_class_sel)
        '''

        #print("predictions")
        #print(predictions)

        probabilities_lista_dividida = np.array_split(probabilities,total_exemplos)
        #selected_classifiers_lista_dividida = np.array_split(selected_classifiers,total_exemplos)
        predictions_lista_dividida = np.array_split(probabilities,total_exemplos)
        #print("probabilities_lista_dividida")
        #print(probabilities_lista_dividida)
        #print("selected_classifiers_lista_dividida")
        #print(selected_classifiers_lista_dividida)
        #print("predictions_lista_dividida")
        #print(predictions_lista_dividida)

        import random
        from random import randrange
        #print("Juntar")
        #print(np.concatenate(probabilities_lista_dividida))
        resultados = []

        from sklearn.neighbors import KNeighborsClassifier
        #neigh = KNeighborsClassifier(n_neighbors=len(y_val2), n_jobs = 1)
        #neigh = KNeighborsClassifier(n_neighbors=5, n_jobs = 1)
        neigh = knn_best_clf_in_grid#pegou as configurações do KNN que vem das configurações do grid
        neigh.fit(X_val2, y_val2)#Treinamento do KNN com as melhores configurações

        X_test_divergente = X_test#.iloc[self.dynamic_index_usage_neighbors_test]
        y_test_divergente = y_test#.iloc[self.dynamic_index_usage_neighbors_test]

        #print(self.dynamic_index_usage_neighbors_test)
        #print("X_test_divergente")
        #print(X_test_divergente)
        '''
        print("self.dynamic_index_usage_neighbors_test")
        print(self.dynamic_index_usage_neighbors_test)
        print("X_test")
        print(X_test)
        print("X_test_divergente")
        print(X_test_divergente)
        print("y_test_divergente")
        print(y_test_divergente)

        print("DO COMITE 1 2 3 4 5")
        print(pool_classifiers[0].predict(X_test_divergente))
        print(pool_classifiers[1].predict(X_test_divergente))
        print(pool_classifiers[2].predict(X_test_divergente))
        print(pool_classifiers[3].predict(X_test_divergente))
        print(pool_classifiers[4].predict(X_test_divergente))
        print("FIM COMITE")
        '''


        #Importação das variaveis globais dos classificadores, quando k=5 ou k=11 esse classsificador é usado quando classificado usando k=3
        global modelMAX 
        global modelSOFT
        global modelHARD 
        global modelMIN 
        global modelG_MEAN 
        global model_Rede_neural 
        global modelRede_neural_soft 
        global modelBorda 
        global modelNaive_bayes 
        global modelPeso_ponderado_comite 
        global modelSum_weight
        global rede_grid
        

        empate_classificadores_no_final = 0
        empate_classificadores_no_inicio = 0
        #row_kneighbors_validation = neigh.kneighbors(X=X_test_divergente, n_neighbors=None, return_distance=False)

        #Pegar X(len val 2) vizinhos de X_teste 
        row_kneighbors_validation = neigh.kneighbors(X=X_test_divergente, n_neighbors=len(y_val2), return_distance=False)

        #print(X_val2)
        #print("Vizinhos do Teste (os vizinhos estão no validation 2)")
        #print(row_kneighbors_validation)
        #print("self.dynamic_index_usage_neighbors_test")
        #print(self.dynamic_index_usage_neighbors_test)

        #lista unica de vizinhos proximos das classes que serão classificadas, os valores estão na base de teste
        row_unique_kneighbors_validation = np.unique(row_kneighbors_validation, return_counts=False)#Dados não mais usados 

        #print("VALORES PARA CLASSIFICAR X")
        #MELHORAR AINDA ESTA CONSUMINDO MUITO
        #X_neighbors_val = X_val2.iloc[row_unique_kneighbors_validation]
        #y_neighbors_val = y_val2.iloc[row_unique_kneighbors_validation]
        X_neighbors_val = X_val2
        y_neighbors_val = y_val2

        #print(X_neighbors_val)
        #print("VALORES PARA CLASSIFICAR Y")
        #print(y_neighbors_val)



        
        #Primeiro teste k=3, segundo teste k=5, terceiro teste k=11
        #O valor de int_k_inserir aponta pra onde será inserido na matriz [0,1,2] cada um representa k=3, k=5, k=7
        if int_k_inserir == 0:
            #print("Treina")

            modelMAX = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="max")
            modelMAX.fit(X_train,y_train )
            #model_predictionsMAX = modelMAX.predict(X_neighbors_val)#predizer dados de treinos
            #proba_MAX =  modelMAX.predict_proba(X_test)
            #print(proba_MAX)
            resultados_max.append(modelMAX.score(X_test, y_test))
            #print(modelMAX.score(X_neighbors_val, y_val2))
            #print(modelMAX.predict(X_test).shape)
            '''
            print("modelMAX.predict(X_test)")
            print(modelMAX.predict(X_test))
            print("modelMAX.predict(X_val2)")
            print(modelMAX.predict(X_val2))
            print("real val 2")
            print(y_val2)
            '''

            #print('TESTE SOFT: ')
            modelSOFT = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="soft")
            modelSOFT.fit(X_train,y_train )
            #model_predictionsSOFT = modelSOFT.predict(X_neighbors_val)#predizer dados de treino
            #proba_SOFT =  modelSOFT.predict_proba(X_test)
            #print(proba_SOFT)
            resultados_soft.append(modelSOFT.score(X_test, y_test))
            #print(modelSOFT.score(X_neighbors_val, y_val2))
            #print(modelSOFT.predict(X_test).shape)


            #print('TESTE HARD: ')
            modelHARD = EnsembleClassifier(estimators=self.estimators, voting='hard', voting_type="soft")#Necessario usar o soft no final para indicar que podemos pegar 
            modelHARD.fit(X_train,y_train )
            #model_predictionsHARD = modelHARD.predict(X_neighbors_val)#predizer dados de treino
            #proba_HARD =  modelHARD.predict_proba(X_test)
            #print("####################################### PROBABILIDADE HARD ##############################################")
            #print(proba_HARD)
            resultados_hard.append(modelHARD.score(X_test, y_test))
            #print(modelHARD.score(X_neighbors_val, y_val2))
            #print(modelHARD.predict(X_test).shape)


            #print('TESTE MIN: ')
            modelMIN = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="min")
            modelMIN.fit(X_train,y_train )
            #model_predictionsMIN = modelMIN.predict(X_neighbors_val)#predizer dados de treinos
            #proba_MIN =  modelMIN.predict_proba(X_test)
            resultados_min.append(modelMIN.score(X_test, y_test))
            #print(modelMIN.score(X_neighbors_val, y_val2))
            #print(modelMIN.predict(X_test).shape)


            #print('TESTE geometric_mean: ')
            modelG_MEAN = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="geometric_mean" )
            modelG_MEAN.fit(X_train,y_train )
            #model_predictionsG_MEAN = modelG_MEAN.predict(X_neighbors_val)#predizer dados de treino
            #proba_G_MEAN =  modelG_MEAN.predict_proba(X_test)
            resultados_geometric_mean.append(modelG_MEAN.score(X_test, y_test))
            #print(modelG_MEAN.score(X_neighbors_val, y_val2))
            #print(modelG_MEAN.predict(X_test).shape)



            #print('TESTE rede_neural: ')
            model_Rede_neural = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="rede_neural")
            model_Rede_neural.fit(X_train,y_train )
            #model_predictionsRede_neural = model_Rede_neural.predict(X_neighbors_val)#predizer dados de trein
            #proba_Rede_neural =  model_Rede_neural.predict_proba(X_test)
            resultados_escolheu_rede_neural.append(model_Rede_neural.score(X_test, y_test))
            #print(model_Rede_neural.score(X_neighbors_val, y_val2))
            #print(model_Rede_neural.predict(X_test).shape)


            #proba_Rede_neural =  model_Rede_neural.predict_proba(X_test)

            #print('TESTE Rede_neural_soft : ')
            modelRede_neural_soft = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="rede_neural_soft")
            modelRede_neural_soft.fit(X_train,y_train )
            #model_predictionsRede_neural_soft = modelRede_neural_soft.predict(X_neighbors_val)#predizer dados de treino
            #proba_Rede_neural_soft =  modelRede_neural_soft.predict_proba(X_test)
            resultados_escolheu_rede_neural_soft.append(modelRede_neural_soft.score(X_test, y_test))
            #print(modelRede_neural_soft.score(X_neighbors_val, y_val2))
            #print(modelRede_neural_soft.predict(X_test).shape)


            #print('TESTE Rede_neural_soft_div: ')
            #modelRede_neural_soft_div = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="rede_neural_soft_div" )
            #modelRede_neural_soft_div.fit(X_train,y_train )
            #model_predictionsRede_neural_soft_div = modelRede_neural_soft_div.predict(X_neighbors_val)#predizer dados de treino
            #proba_Rede_neural_soft_div =  modelRede_neural_soft_div.predict_proba(X_test)
            #resultados_escolheu_rede_neural_soft_div.append(modelRede_neural_soft_div.score(X_test, y_test))
            #print(modelRede_neural_soft_div.score(X_neighbors_val, y_val2))



            #print('TESTE borda: ')
            modelBorda = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="borda" )
            modelBorda.fit(X_train,y_train )
            #model_predictionsBorda = modelBorda.predict(X_neighbors_val)#predizer dados de treino
            #proba_Borda =  modelBorda.predict_proba(X_test)
            resultados_escolheu_borda.append(modelBorda.score(X_test, y_test))
            #print(modelBorda.score(X_neighbors_val, y_val2))
            #print(modelBorda.predict(X_test).shape)


            #print('TESTE naive_bayes: ')
            modelNaive_bayes = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="naive_bayes" )
            modelNaive_bayes.fit(X_train,y_train )
            #model_predictionsNaive_bayes = modelNaive_bayes.predict(X_neighbors_val)#predizer dados de treino
            #proba_Naive_bayes =  modelNaive_bayes.predict_proba(X_test)
            resultados_escolheu_naive_bayes.append(modelNaive_bayes.score(X_test, y_test))
            #print(modelNaive_bayes.score(X_neighbors_val, y_val2))
            #print(modelNaive_bayes.predict(X_test).shape)


            #####TIRAR
            #print('TESTE Peso_ponderado_classe: ')
            #print("50")
            #modelPeso_ponderado_classe = MyKnoraE(pool_classifiers=pool_classifiers, k=k_dynamic_combination, voting_type="peso_ponderado_classe_cada_amostra_ajustado_0a1" )
            #print("51")
            #modelPeso_ponderado_classe.fit(X_train,y_train )
            #print("52")
            #model_predictionsPeso_ponderado_classe = modelPeso_ponderado_classe.predict(X_neighbors_val)#predizer dados de treino
            #print("53")
            #resultados_peso_ponderado_classe_cada_amostra_ajustado_0a1.append(modelPeso_ponderado_classe.score(X_test, y_test))
            #print("54")
            #proba_Peso_ponderado_classe =  modelPeso_ponderado_classe.predict_proba(X_test)
            #print("55")
            #print(modelPeso_ponderado_classe.score(X_neighbors_val, y_val2))


            #print('TESTE peso_ponderado_comite: ')
            modelPeso_ponderado_comite = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1" )
            #print("60")
            modelPeso_ponderado_comite.fit(X_train,y_train )
            #print("61")
            #model_predictionsPeso_ponderado_comite = modelPeso_ponderado_comite.predict(X_neighbors_val)#predizer dados de treino
            #print("62")
            #proba_Peso_ponderado_comite =  modelPeso_ponderado_comite.predict_proba(X_test)
            #print("63")
            resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1.append(modelPeso_ponderado_comite.score(X_test, y_test))
            #print("64")
            #print(modelPeso_ponderado_comite.score(X_neighbors_val, y_val2))
            #print(modelPeso_ponderado_comite.predict(X_test).shape)



            #print('TESTE sum_weight: ')
            modelSum_weight = EnsembleClassifier(estimators=self.estimators, voting='soft', voting_type="sum_weight_0a1_votes_per_class" )
            #print("71")
            modelSum_weight.fit(X_train,y_train )
            #print("72")
            #model_predictionsSum_weight = modelSum_weight.predict(X_neighbors_val)#predizer dados de treino
            #print("73")
            #proba_Sum_weight =  modelSum_weight.predict_proba(X_test)
            #print("74")
            resultados_sum_weight_0a1_votes_per_class.append(modelSum_weight.score(X_test, y_test))
            #print("75")
            #print(modelSum_weight.score(X_neighbors_val, y_val2))
            #print(modelSum_weight.predict(X_test).shape)

        #Desabilita o rede grid, assim não realizara mais o gid da MLP
        rede_grid = False


        #print("COMBINAÇÂO DINAMICA")





        #print('RESULTADO TESTE VAL REAL: ')
        #print(y_neighbors_val)
        #print('RESULTADO TESTE VAL SOFT: ')
        #print(model_predictionsSOFT)



        #Só executa uma vez em cada iteração mesmo com k=3, k=5 e k=11
        if int_k_inserir == 0:
            resultado_MAX =  modelMAX.predict(X_test)
            resultado_SOFT =  modelSOFT.predict(X_test)
            resultado_HARD =  modelHARD.predict(X_test)
            resultado_MIN =  modelMIN.predict(X_test)
            resultado_G_MEAN =  modelG_MEAN.predict(X_test)
            resultado_Rede_neural =  model_Rede_neural.predict(X_test)
            resultado_Rede_neural_soft =  modelRede_neural_soft.predict(X_test)
            #resultado_Rede_neural_soft_div =  modelRede_neural_soft_div.predict(X_test)
            resultado_Borda =  modelBorda.predict(X_test)
            resultado_Naive_bayes =  modelNaive_bayes.predict(X_test)
            resultado_Peso_ponderado_comite =  modelPeso_ponderado_comite.predict(X_test)
            resultado_Sum_weight =  modelSum_weight.predict(X_test)
            #print("resultado_MAX")
            #print(resultado_MAX)
            #print("resultado_HARD")
            #print(resultado_HARD)
            #print("resultado_SOFT")
            #print(resultado_SOFT)
            #print("resultado_MIN")
            #print(resultado_MIN)
            #print("resultado_Rede_neural_soft")
            #print(resultado_Rede_neural_soft)

        #Só executa uma vez em cada iteração mesmo com k=3, k=5 e k=11
        if int_k_inserir == 0:
            acertos_maximo = 0
            for i in range(len(y_test)):
                #if  (resultado_SOFT[i] == y_test.iloc[i]) or (resultado_HARD[i] == y_test.iloc[i])  or (resultado_Rede_neural_soft[i] == y_test.iloc[i]) :
                #if (resultado_MAX[i] == y_test.iloc[i]) or (resultado_SOFT[i] == y_test.iloc[i]) or (resultado_HARD[i] == y_test.iloc[i]) or (resultado_MIN[i] == y_test.iloc[i]) or (resultado_G_MEAN[i] == y_test.iloc[i]) or (resultado_Rede_neural[i] == y_test.iloc[i]) or (resultado_Rede_neural_soft[i] == y_test.iloc[i]) or (resultado_Borda[i] == y_test.iloc[i]) or (resultado_Naive_bayes[i] == y_test.iloc[i]) or (resultado_Peso_ponderado_comite[i] == y_test.iloc[i]) or (resultado_Sum_weight[i] == y_test.iloc[i]):
                if (resultado_SOFT[i] == y_test.iloc[i]) or (resultado_HARD[i] == y_test.iloc[i]) or (resultado_G_MEAN[i] == y_test.iloc[i]) or (resultado_Rede_neural[i] == y_test.iloc[i]) or (resultado_Rede_neural_soft[i] == y_test.iloc[i]) or (resultado_Borda[i] == y_test.iloc[i]) or (resultado_Naive_bayes[i] == y_test.iloc[i]) :
                    acertos_maximo  = acertos_maximo + 1


            resultados_maximo_na_combinacao.append(acertos_maximo / len(y_test))




        #print(len(self.dynamic_index_usage_neighbors_test))




        '''
        print("vizinhos_val")
        print(vizinhos_val)
        print("row_kneighbors_validation")
        print(row_kneighbors_validation)
        print("row_kneighbors_validation.shape")
        print(row_kneighbors_validation.shape)
        print("self.dynamic_index_usage_neighbors_test")
        print(self.dynamic_index_usage_neighbors_test.shape)
        '''
        #print("MOSTRAR RESULTADO")
        
        lin, col = row_kneighbors_validation.shape
        resultados = []
        #print("MAX SOFT HARD MIN G_MEAN")
        
        #Quantidade de escolha de cada método em numeros absolutos
        escolheu_MAX = 0
        escolheu_SOFT = 0
        escolheu_HARD = 0
        escolheu_MIN = 0
        escolheu_G_MEAN = 0
        #escolheu_peso_ponderado_classe = 0
        escolheu_peso_ponderado_comite = 0
        escolheu_sum_weight = 0
        #escolheu_sum_weight_line = 0
        escolheu_rede_neural = 0
        escolheu_rede_neural_soft = 0
        escolheu_rede_neural_soft_div = 0
        escolheu_borda = 0
        escolheu_naive_bayes = 0
        escolheu_compartilhado = 0
        #print("Os desacordo")
        #print(self.dynamic_index_usage_neighbors_test)

        #PREVISÃO DA BASE DE VALIDAÇÃO 2, A PREVISÃO SERA USADA PARA VERIFICAR A REGIÃO DE COMPETENCIA 
        #pMAX = modelMAX.predict(X_val2)
        pSOFT = modelSOFT.predict(X_val2)
        pHARD = modelHARD.predict(X_val2)
        #pMIN = modelMIN.predict(X_val2)
        pG_MEAN = modelG_MEAN.predict(X_val2)
        #pPeso_ponderado_comite = modelPeso_ponderado_comite.predict(X_val2)
        #pSum_weight = modelSum_weight.predict(X_val2)
        pRede_neural = model_Rede_neural.predict(X_val2)
        pRede_neural_soft = modelRede_neural_soft.predict(X_val2)
        #pRede_neural_soft_class = modelRede_neural_soft_div.predict(X_val2)
        pBorda = modelBorda.predict(X_val2)
        pNaive_bayes = modelNaive_bayes.predict(X_val2)

        # I  de 0 a x de quantos precisam do comitê
        #for i in range(len(self.dynamic_index_usage_neighbors_test)):
        #PARA teste DE primeiro exemplo de teste ATÉ ultimo exemplo de teste FAÇA:
        for i in range(len(X_test)):

            numero_colunas = 0
            #Habilitado para continuar na busca do mais competente
            continuar_MAX = False#Desabilitado em escolhas
            continuar_SOFT = True
            continuar_HARD = True
            continuar_MIN = False#Desabilitado em escolhas
            continuar_G_MEAN = True
            #continuar_peso_ponderado_classe = True
            continuar_peso_ponderado_comite = False#Desabilitado em escolhas
            continuar_sum_weight = False#Desabilitado em escolhas
            #continuar_sum_weight_line = True
            continuar_rede_neural = True
            continuar_rede_neural_soft = True
            continuar_rede_neural_soft_div = False#Desabilitado em escolhas
            continuar_borda = True
            continuar_naive_bayes = True
            #print("######## VIZINHO "+str(i)+ "########")

            #Acertos da base de validação 2 para  cada tipo de comitê
            acertos_MAX = -1 #Desabilitado em escolhas
            acertos_SOFT = 0
            acertos_HARD = 0
            acertos_MIN = -1 #Desabilitado em escolhas
            acertos_G_MEAN = 0
            acertos_peso_ponderado_comite = -1 #Desabilitado em escolhas
            acertos_sum_weight = -1 #Desabilitado em escolhas
            acertos_rede_neural = 0
            acertos_rede_neural_soft = 0
            acertos_rede_neural_soft_div = -1 #Desabilitado em escolhas
            acertos_borda = 0
            acertos_naive_bayes = 0

            #print("self.dynamic_index_usage_neighbors_test[i]")
            #print(self.dynamic_index_usage_neighbors_test[i])
            #a_classificar_X_teste = X_test.iloc[self.dynamic_index_usage_neighbors_test[i]]
            #a_classificar_X_teste = X_test.iloc[X_test[i]]

            #Primeiro a classificar, segundo, terceiro #i... O iloc pega o elemento que se quer do DataFrame
            a_classificar_X_teste = X_test.iloc[i]#i
            
            #print(a_classificar_X_teste)
            #print("Classificar")
            #print(a_classificar_X_teste)

            qt_vizinhos_classificados = 0
            # primeiro, segundo divergente  dos vizinhos de teste
            #print("OS vizinhos")
            #print(row_kneighbors_validation[i])

            #PARA vizinho DE Teste ATÉ ultimo vizinho de teste FAÇA:
            #Roda dos vizinhos mais proximos até os mais longes da amostra qeu se quer classificar #i
            for vizinho in row_kneighbors_validation[i]:
                #print("vizinho ######################################################################")
                #print(vizinho)
                
                #Incrementa quantos vizinhos foram testados 
                qt_vizinhos_classificados = qt_vizinhos_classificados + 1

                # Classe REAL do exemplo de validação 2, dos vizinhos testados. 
                real_y_val2_classificar = y_val2.iloc[vizinho]
                '''
                preveuMAX = modelMAX.predict([X_val2.iloc[vizinho]])
                preveuSOFT = modelSOFT.predict([X_val2.iloc[vizinho]])
                preveuHARD = modelHARD.predict([X_val2.iloc[vizinho]])
                preveuMIN = modelMIN.predict([X_val2.iloc[vizinho]])
                preveuG_MEAN = modelG_MEAN.predict([X_val2.iloc[vizinho]])
                preveuPeso_ponderado_comite = modelPeso_ponderado_comite.predict([X_val2.iloc[vizinho]])
                preveuSum_weight = modelSum_weight.predict([X_val2.iloc[vizinho]])
                preveuRede_neural = model_Rede_neural.predict([X_val2.iloc[vizinho]])
                preveuRede_neural_soft = modelRede_neural_soft.predict([X_val2.iloc[vizinho]])
                preveuRede_neural_soft_class = modelRede_neural_soft_div.predict([X_val2.iloc[vizinho]])
                preveuBorda = modelBorda.predict([X_val2.iloc[vizinho]])
                preveuNaive_bayes = modelNaive_bayes.predict([X_val2.iloc[vizinho]])
                '''

                # Clase PREVISTA do exemplo de validação 2, dos vizinhos testados. Para cada comitê de classificador
                #preveuMAX = pMAX[vizinho]
                preveuSOFT = pSOFT[vizinho]
                preveuHARD = pHARD[vizinho]
                #preveuMIN = pMIN[vizinho]
                preveuG_MEAN = pG_MEAN[vizinho]
                #preveuPeso_ponderado_comite = pPeso_ponderado_comite[vizinho]
                #preveuSum_weight = pSum_weight[vizinho]
                preveuRede_neural = pRede_neural[vizinho]
                preveuRede_neural_soft = pRede_neural_soft[vizinho]
                #preveuRede_neural_soft_class = pRede_neural_soft_class[vizinho]
                preveuBorda = pBorda[vizinho]
                
                preveuNaive_bayes = pNaive_bayes[vizinho]

                outroPreveuSOFT = modelSOFT.predict([X_val2.iloc[vizinho]])
                if (preveuSOFT != outroPreveuSOFT[0]):
                    print("###################### ############################### #################################### ############## ALGO DE ERRADO NO CODIGO HARD  ###################### ############################### #################################### ############## ")


                outroPreveuNaive = modelNaive_bayes.predict([X_val2.iloc[vizinho]])
                if (preveuNaive_bayes != outroPreveuNaive[0]):
                    print("###################### ############################### #################################### ############## ALGO DE ERRADO NO CODIGO NAIVE  ###################### ############################### #################################### ############## ")


                #print("Real "+str(real_y_val2_classificar)+" MAX "+str(preveuMAX)+" HARD "+str(preveuHARD)+" SOFT "+str(preveuSOFT)+" MAX "+str(preveuMAX)+" MIN "+str(preveuMIN)+" G_MEAN "+str(preveuG_MEAN)+" Peso_ponderado_comite "+str(preveuPeso_ponderado_comite)+" Sum_weight "+str(preveuSum_weight)+" Rede_neural "+str(preveuRede_neural)+" Rede_neural_soft "+str(preveuRede_neural_soft)+" Rede_neural_soft_class "+str(preveuRede_neural_soft_class)+" Borda "+str(preveuBorda)+" Naive_bayes "+str(preveuNaive_bayes))
                #print( str(preveuMAX)+ " - "+ str(modelMAX.predict([X_val2.iloc[vizinho]]))   )

                #Aqui são testados quem dos classificadores acertou o vizinho (VAL2)
                #Quem acertou tem o valor de acerto incrementado
                #Quem está desabilitado não é mais testado
                '''
                if(continuar_MAX and preveuMAX == real_y_val2_classificar):
                    #print("acertou max")
                    acertos_MAX = acertos_MAX + 1
                '''
                if( continuar_SOFT and  preveuSOFT == real_y_val2_classificar):
                    #print("acertou soft")
                    acertos_SOFT = acertos_SOFT + 1

                if( continuar_HARD and  preveuHARD == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_HARD = acertos_HARD + 1
                '''
                if( continuar_MIN and  preveuMIN == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_MIN = acertos_MIN + 1
                '''
                if( continuar_G_MEAN and  preveuG_MEAN == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_G_MEAN = acertos_G_MEAN + 1
                '''
                if( continuar_peso_ponderado_comite and  preveuPeso_ponderado_comite == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_peso_ponderado_comite = acertos_peso_ponderado_comite + 1

                if( continuar_sum_weight and  preveuSum_weight == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_sum_weight = acertos_sum_weight + 1
                '''
                if( continuar_rede_neural and  preveuRede_neural == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_rede_neural = acertos_rede_neural + 1
                if( continuar_rede_neural_soft and  preveuRede_neural_soft == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_rede_neural_soft = acertos_rede_neural_soft + 1
                '''
                if( continuar_rede_neural_soft_div and  preveuRede_neural_soft_class == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_rede_neural_soft_div = acertos_rede_neural_soft_div + 1
                '''
                if( continuar_borda and  preveuBorda == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_borda = acertos_borda + 1

                if( continuar_naive_bayes and  preveuNaive_bayes == real_y_val2_classificar):
                    #print("acertou HARD")
                    acertos_naive_bayes = acertos_naive_bayes + 1



                #ESCOLHER (se) QUEM MAIS ACERTOU
                #Quando o valor da região de competencia for alcancado(ViziVal) 
                if qt_vizinhos_classificados >= vizinhos_val:
                    '''
                    if acertos_MAX > max( acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        #print("############################# ESCOLHEU MAX ############################")
                        resultados.append(modelMAX.predict_proba([a_classificar_X_teste]))
                        escolheu_MAX = escolheu_MAX + 1
                        break
                    '''
                    #print("acertos_HARD: " +str(acertos_HARD)+", acertos_SOFT: " +str(acertos_SOFT)+", acertos_MAX: " +str(acertos_MAX)+", acertos_MIN: " +str(acertos_MIN)+", acertos_G_MEAN: " +str(acertos_G_MEAN)+", acertos_peso_ponderado_comite: " +str(acertos_peso_ponderado_comite)+", acertos_sum_weight: " +str(acertos_sum_weight)+", acertos_rede_neural: " +str(acertos_rede_neural)+", acertos_rede_neural_soft: " +str(acertos_rede_neural_soft)+", acertos_rede_neural_soft_div: " +str(acertos_rede_neural_soft_div)+", acertos_borda: " +str(acertos_borda)+", acertos_naive_bayes: " +str(acertos_naive_bayes))

                    #Se esse comite de classificador foi o que mais acertou ele é escolhido e partido pra proxima iteração, no caso proximo teste      
                    if acertos_HARD > max(acertos_MAX, acertos_SOFT, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        #print("############################# ESCOLHEU HARD ############################")
                        resultados.append(modelHARD.predict_proba([a_classificar_X_teste]))
                        escolheu_HARD = escolheu_HARD + 1
                        #print("VM")
                        #print(modelHARD.predict_proba([a_classificar_X_teste]))
                        break
                    if acertos_SOFT > max(acertos_MAX, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        #print("############################# ESCOLHEU SOFT ############################")
                        resultados.append(modelSOFT.predict_proba([a_classificar_X_teste]))
                        escolheu_SOFT = escolheu_SOFT + 1
                        #print("Media")
                        #print(modelSOFT.predict_proba([a_classificar_X_teste]))
                        break
                    '''
                    if acertos_MIN > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        #print("############################# ESCOLHEU MIN ############################")
                        resultados.append(modelMIN.predict_proba([a_classificar_X_teste]))
                        escolheu_MIN = escolheu_MIN + 1
                        break
                    '''
                    if acertos_G_MEAN > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        resultados.append(modelG_MEAN.predict_proba([a_classificar_X_teste]))
                        escolheu_G_MEAN = escolheu_G_MEAN + 1
                        #print("g mean")
                        #print(modelG_MEAN.predict_proba([a_classificar_X_teste]))
                        break
                    '''
                    if acertos_peso_ponderado_comite > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        resultados.append(modelPeso_ponderado_comite.predict_proba([a_classificar_X_teste]))
                        escolheu_peso_ponderado_comite = escolheu_peso_ponderado_comite + 1
                        break
                    if acertos_sum_weight > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        resultados.append(modelSum_weight.predict_proba([a_classificar_X_teste]))
                        escolheu_sum_weight = escolheu_sum_weight + 1
                        break
                    '''
                    if acertos_rede_neural > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        resultados.append(model_Rede_neural.predict_proba([a_classificar_X_teste]))
                        escolheu_rede_neural = escolheu_rede_neural + 1
                        #print("rna")
                        #print(a_classificar_X_teste)
                        #print(model_Rede_neural.predict_proba([a_classificar_X_teste]))
                        break

                        #print("a_classificar_X_teste")
                        #print(a_classificar_X_teste)
                    if acertos_rede_neural_soft > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        #print("############################# ESCOLHEU RNA ############################")
                        resultados.append(modelRede_neural_soft.predict_proba([a_classificar_X_teste]))
                        escolheu_rede_neural_soft = escolheu_rede_neural_soft + 1
                        #print("rna soft")
                        #print(modelRede_neural_soft.predict_proba([a_classificar_X_teste]))
                        break
                    '''
                    if acertos_rede_neural_soft_div > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_borda, acertos_naive_bayes ):
                        resultados.append(modelRede_neural_soft_div.predict_proba([a_classificar_X_teste]))
                        escolheu_rede_neural_soft_div = escolheu_rede_neural_soft_div + 1
                        break
                    '''
                    if acertos_borda > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_naive_bayes ):
                        resultados.append(modelBorda.predict_proba([a_classificar_X_teste]))
                        escolheu_borda = escolheu_borda + 1
                        #print("Borda")
                        #print(modelBorda.predict_proba([a_classificar_X_teste]))
                        break
                    if acertos_naive_bayes > max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda ):
                        resultados.append(modelNaive_bayes.predict_proba([a_classificar_X_teste]))
                        escolheu_naive_bayes = escolheu_naive_bayes + 1
                        #print("Naive")
                        #print(modelNaive_bayes.predict_proba([a_classificar_X_teste]))
                        break

                    #Se ninguem tiver acertado mais que os outros então é incrementado o valor de empate na região inicial
                    #Se alguem tiver sido ecolhido anteriormente então nem chega nessa parte do codigo 
                    if qt_vizinhos_classificados == vizinhos_val:
                        empate_classificadores_no_inicio = empate_classificadores_no_inicio + 1 



                    #Se ainda tiver com empate então
                    #DESABILITA QUEM ERROU o teste, caso alguém tenha acertado
                    '''
                    if acertos_MAX < max( acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_MAX = False
                    '''
                    if acertos_HARD < max(acertos_MAX, acertos_SOFT, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_HARD = False
                    if acertos_SOFT < max(acertos_MAX, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_SOFT = False
                    '''
                    if acertos_MIN < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_MIN = False
                    '''
                    if acertos_G_MEAN < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_G_MEAN = False
                    '''
                    if acertos_peso_ponderado_comite < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_peso_ponderado_comite = False
                    if acertos_sum_weight < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_sum_weight = False
                    '''
                    if acertos_rede_neural < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_rede_neural = False

                    if acertos_rede_neural_soft < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft_div, acertos_borda, acertos_naive_bayes ):
                        continuar_rede_neural_soft = False
                    '''
                    if acertos_rede_neural_soft_div < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_borda, acertos_naive_bayes ):
                        continuar_rede_neural_soft_div = False
                    '''
                    if acertos_borda < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_naive_bayes ):
                        continuar_borda = False
                    if acertos_naive_bayes < max(acertos_MAX, acertos_SOFT, acertos_HARD, acertos_MIN, acertos_G_MEAN, acertos_peso_ponderado_comite, acertos_sum_weight, acertos_rede_neural, acertos_rede_neural_soft, acertos_rede_neural_soft_div, acertos_borda ):
                        continuar_naive_bayes = False





                    #SE RODOU METADE DA REGIÃO DE COMPETENCIA
                    if qt_vizinhos_classificados >= (len(y_val2)/2):
                        empate_classificadores_no_final = empate_classificadores_no_final + 1
                        escolheu_compartilhado = escolheu_compartilhado + 1
                        media_desse_exemplo = []

                        #print( str(continuar_HARD) + " - "+ str(continuar_SOFT) +" - "+ str(continuar_G_MEAN) +" - "+ str(continuar_rede_neural) +" - "+ str(continuar_rede_neural_soft) +" - "+ str(continuar_borda) +" - "+ str(continuar_naive_bayes) )
                        

                        #FUNDIR POR MEDIA OS CLASSIFICADORES HABILITADOS
                        if continuar_HARD == True:
                            #print(modelHARD.predict_proba([a_classificar_X_teste]))
                            media_desse_exemplo.append(modelHARD.predict_proba([a_classificar_X_teste]))
                            #escolheu_HARD = escolheu_HARD + 1
                            #print("############################# ESCOLHEU HARD ############################")
                            #print(modelHARD.predict_proba([a_classificar_X_teste]))
                            #break
                        if continuar_SOFT == True:
                            #print(modelSOFT.predict_proba([a_classificar_X_teste]))
                            media_desse_exemplo.append(modelSOFT.predict_proba([a_classificar_X_teste]))
                            #escolheu_SOFT = escolheu_SOFT + 1
                            #print("############################# ESCOLHEU SOFT ############################")
                            #print(modelSOFT.predict_proba([a_classificar_X_teste]))
                            #break
                        if continuar_G_MEAN == True:
                            #print(modelG_MEAN.predict_proba([a_classificar_X_teste]))
                            media_desse_exemplo.append(modelG_MEAN.predict_proba([a_classificar_X_teste]))
                            #escolheu_G_MEAN = escolheu_G_MEAN + 1
                            #print("########## EScolheu G mean ################")
                            #print(modelG_MEAN.predict_proba([a_classificar_X_teste]))
                            #break
                        if continuar_rede_neural == True:
                            #print(model_Rede_neural.predict_proba([a_classificar_X_teste]))
                            media_desse_exemplo.append(model_Rede_neural.predict_proba([a_classificar_X_teste]))
                            #escolheu_rede_neural = escolheu_rede_neural + 1
                            #print("############### Escolheu RNA ###############")
                            #print(a_classificar_X_teste)
                            #print(model_Rede_neural.predict_proba([a_classificar_X_teste]))
                            #break

                        if continuar_rede_neural_soft == True:
                            #print(modelRede_neural_soft.predict_proba([a_classificar_X_teste]))
                            media_desse_exemplo.append(modelRede_neural_soft.predict_proba([a_classificar_X_teste]))
                            #print("############################# ESCOLHEU RNA SOFT ############################")
                            #escolheu_rede_neural_soft = escolheu_rede_neural_soft + 1
                            #print(modelRede_neural_soft.predict_proba([a_classificar_X_teste]))
                            #break
                        if continuar_borda == True:
                            #print(modelBorda.predict_proba([a_classificar_X_teste]))
                            media_desse_exemplo.append(modelBorda.predict_proba([a_classificar_X_teste]))
                            #escolheu_borda = escolheu_borda + 1
                            #print("escolher Borda ###")
                            #print(modelBorda.predict_proba([a_classificar_X_teste]))
                            #break
                        if continuar_naive_bayes == True:
                            #print(modelNaive_bayes.predict_proba([a_classificar_X_teste]))
                            media_desse_exemplo.append(modelNaive_bayes.predict_proba([a_classificar_X_teste]))
                            #escolheu_naive_bayes = escolheu_naive_bayes + 1
                            #print("Escolher Naive #######")
                            #print(modelNaive_bayes.predict_proba([a_classificar_X_teste]))
                            #print("empate")
                            #break

                        #escolheu_compartilhado = escolheu_compartilhado + 1

                        #print("Resultado")
                        #print(media_desse_exemplo)
                        #print("Media desses clas")
                        #print(np.average(media_desse_exemplo, axis=0))
                        resultados.append(np.average(media_desse_exemplo, axis=0))
                        break

                        ########################################## FIM ###########################################################################    
                    
                        #Não pode chegar aqui mais
                        print("Erro por ter rodado d+")
      



        #print("RESULTADOS")
        #print(resultados)

        #print(np.concatenate(resultados))
        #print(resultados)
        #return resultados
        #print("Resultado função juntado")
        #print(np.concatenate(resultados))
        #print("1")


        #Porcentagem de escolha de cada comite na fusão dinamica
        #escolheu_TOTAL = sum([escolheu_borda, escolheu_naive_bayes, escolheu_peso_ponderado_comite,escolheu_rede_neural_soft_div, escolheu_MAX,escolheu_SOFT,escolheu_HARD, escolheu_MIN, escolheu_G_MEAN, escolheu_sum_weight, escolheu_rede_neural,escolheu_rede_neural_soft, ])
        escolheu_TOTAL = sum([escolheu_compartilhado, escolheu_borda, escolheu_naive_bayes,escolheu_SOFT,escolheu_HARD, escolheu_G_MEAN, escolheu_rede_neural,escolheu_rede_neural_soft, ])
        #print("escolheu TOTAL")
        #print(escolheu_TOTAL)

        #Inserido no local da matriz de cada porcentagem de comitê [int_k_inserir]
        #porcentagem_escolheu_MAX[int_k_inserir].append((escolheu_MAX / escolheu_TOTAL) * 100)
        porcentagem_escolheu_SOFT[int_k_inserir].append((escolheu_SOFT / escolheu_TOTAL) * 100)
        porcentagem_escolheu_HARD[int_k_inserir].append((escolheu_HARD / escolheu_TOTAL) * 100)
        #porcentagem_escolheu_MIN[int_k_inserir].append((escolheu_MIN / escolheu_TOTAL) * 100)
        porcentagem_escolheu_G_MEAN[int_k_inserir].append((escolheu_G_MEAN / escolheu_TOTAL) * 100)

        #porcentagem_escolheu_sum_weight[int_k_inserir].append((escolheu_sum_weight / escolheu_TOTAL) * 100)
        #porcentagem_escolheu_peso_ponderado_comite[int_k_inserir].append((escolheu_peso_ponderado_comite / escolheu_TOTAL) * 100)

        porcentagem_escolheu_rede_neural[int_k_inserir].append((escolheu_rede_neural / escolheu_TOTAL) * 100)
        porcentagem_escolheu_rede_neural_soft[int_k_inserir].append((escolheu_rede_neural_soft / escolheu_TOTAL) * 100)
        #porcentagem_escolheu_rede_neural_soft_div.append((escolheu_rede_neural_soft_div / escolheu_TOTAL) * 100)
        porcentagem_escolheu_borda[int_k_inserir].append((escolheu_borda / escolheu_TOTAL) * 100)
        porcentagem_escolheu_naive_bayes[int_k_inserir].append((escolheu_naive_bayes / escolheu_TOTAL) * 100)

        porcentagem_escolha_compartilhada[int_k_inserir].append((escolheu_compartilhado / escolheu_TOTAL) * 100)

        #Não usado
        quantidade_exemplos_divergencia.append(len(X_test_divergente))#Não usado
        #quantidade_classificadores_selecionados.append(qnt_class_sel)
        #print("Q exemplos divergentes")
        #print(len(X_test_divergente))

        #print("Q class selec")
        #print(qnt_class_sel)

        '''
        print(escolheu_MAX)
        print(escolheu_SOFT)
        print(escolheu_HARD)
        print(escolheu_MIN)
        print(escolheu_G_MEAN)
        '''
        #print("empate_classificadores_no_inicio: "+str(empate_classificadores_no_inicio))
        #print("empate_classificadores_no_final: "+str(empate_classificadores_no_final))

        #Inserido no local da matriz de empate [int_k_inserir]
        geral_empate_classificadores_no_inicio[int_k_inserir].append(empate_classificadores_no_inicio)
        geral_empate_classificadores_no_final[int_k_inserir].append(empate_classificadores_no_final)
        #print("Devolve")
        #print (np.concatenate(resultados))
        #print (np.concatenate(resultados).shape)

        return np.concatenate(resultados)#Devolve os resultados



    def rede_neural_soft(self, probabilities):#Devolve as probabilidades mas remove dos classificadores não competentes
        global classificador_Rede_neural
        global rede_grid
        global rede_soft_grid_estimador
        #lista_treinamento = []
        #modelSOFT = EnsembleClassifier(estimators=self.estimators, voting="soft" )
        #MONTAR DADOS DE TREINO INICIO
        if int_k_inserir == 0:
            try:
                #print("1")
                modelSOFT = EnsembleClassifier(estimators=self.estimators, voting="soft", voting_type="max" )
                #print("2")

                #print("3")
                modelSOFT.fit(X_train,y_train )
                #print("4")
                model_predictionsSOFT = modelSOFT.predict(X_val)#predizer dados de treino

                #print("modelSOFT.probabilidades_classificadores")
                #print(modelSOFT.probabilidades_classificadores)
                xxx, yyy, zzz = modelSOFT.probabilidades_classificadores.shape
                lista_treinamento = []
                for x in range(xxx):
                    #for y in range(yyy):
                    #print(modelSOFT.probabilidades_classificadores[x, y])
                    sub = modelSOFT.probabilidades_classificadores[x].flatten().tolist()
                    for z in range(len(sub)):
                        if sub[z] is None:
                            #sub[z] = 1/len(self.classes_)
                            sub[z] = -1
                    lista_treinamento.append(sub)
                    #De 3D para 2D
            except:
                #print("Deu algum erro")
                #DEU erro ao importar os dados dentreino da rede pq esta vazio o treinamento
                pass
            #print("lista_treinamento")
            #print(lista_treinamento)
        #MONTAR DADOS DE TREINO FIM

        #MONTAR DADOS DE TESTE INICIO
        from sklearn.neural_network import MLPClassifier

        xxx, yyy, zzz = probabilities.shape
        masked_proba2D = []
        for x in range(xxx):
            sub = probabilities[x].flatten().tolist()
            for z in range(len(sub)):
                if sub[z] is None:
                    #sub[z] = 1/len(self.classes_)
                    sub[z] = -1
            masked_proba2D.append(sub)

            #De 3D para 2D
        #print("10")
        #MONTAR DADOS DE TESTE FIM

        #Criar classificador
        #clf = MLPClassifier(hidden_layer_sizes=(200,), solver='adam', alpha=0.0001, max_iter=1000, random_state=1)
        #print("11")
        #clf = MLPClassifier(hidden_layer_sizes=(10,), solver='adam', alpha=0.01, max_iter=10, random_state=1)
        parameters = {'early_stopping':[False, True], 'validation_fraction': [0.005, 0.1, 0.2], 'max_iter':[500], 'hidden_layer_sizes':[(100,), (50,), (200,), (300,)], 'learning_rate':['constant', 'invscaling', 'adaptive'], 'solver':['sgd']}    
        clf = GridSearchCV(estimator = MLPClassifier(), param_grid = parameters, refit=True, cv = 5)
        try:
            #print("122")
            #clf.fit(modelSOFT.previsoes_real)
            #clf.fit(lista_treinamento, modelSOFT.previsoes_real)
            if int_k_inserir == 0:
                #print("123")
                if rede_grid:
                    #print("Rede soft")
                    clf.fit(lista_treinamento, y_val)#O erro é aqui
                    #print(clf.best_params_)
                    rede_soft_grid_estimador = clf.best_estimator_
                    clf = rede_soft_grid_estimador
                    #print("0012")
                else:
                    #print("12123")
                    clf = rede_soft_grid_estimador
                    clf.fit(lista_treinamento, y_val)
                    #print("1253")
                classificador_Rede_neural = clf
                #print("12245")
            else:
                clf = classificador_Rede_neural
                #print("1200")
                
            #print("13")
            predicted_pesos  = clf.predict_proba(masked_proba2D)
            #print("1###4")
        except:
            predicted_pesos = np.array([[0]])
            #print("15")

        #print("masked_proba2D")
        #print(masked_proba2D)
        #print("clf.predict(masked_proba2D)")
        #print(clf.predict(masked_proba2D))
        #print("Probabilidade")
        #print(predicted_pesos)


        #clf.predict([[2., 2.], [-1., -2.]])



        '''
        x, y, = predicted_pesos.shape
        if(len(self.classes_) > y):
            #print("Ajuste rede 2")
            df_dados_x = []
            #print(X_val)
            #print("fim")
            for amostra in range(len(X_val)):
                #print(X_val.iloc[amostra:amostra+1,:])
                sub = np.array([])
                for pool in self.pool_classifiers:
                    p = pool.predict_proba(X_val.iloc[amostra:amostra+1,:]).flatten().tolist()
                    sub = np.append(sub, p)
                    #print("saiu")
                dados_amostra = sub.flatten().tolist()
                #aaa, bbb, = sub.shape
                #fo a in range(aaa):
                #    for b in range(bbb):
                #        dados_amostra.append(sub[a,b])
                #print(dados_amostra)
                df_dados_x.append(dados_amostra)

            #print(df_dados_x)
            clf = MLPClassifier(hidden_layer_sizes=(200,), solver='adam', alpha=0.0001, max_iter=1000, random_state=1)
            clf.fit(df_dados_x, y_val)
            predicted_pesos  = clf.predict_proba(masked_proba2D)
        '''

        #x, y, = predicted_pesos.shape
        #if(len(self.classes_) > y):
        #    #print(y_val)
        #    predicted_pesos = self._mask_proba(pbb, slclsf)

        return predicted_pesos

    def rede_neural_class(self, probabilities):#Devolve as probabilidades mas remove dos classificadores não competentes
        global classificador_neural_soft
        global classificador_neural_softAUX
        global rede_grid
        global rede_grid_estimador
        #PEGAR OS RECURSOS DOS DADOS DE TREINO EX [0 1 0 1 0 1] 
        if int_k_inserir == 0:
            modelHARD = EnsembleClassifier(estimators=self.estimators, voting="hard", voting_type="hard" )
            modelHARD.fit(X_train,y_train )
            model_predictionsHARD = modelHARD.predict(X_val)#predizer dados de treino
        #proba_HARD =  modelHARD.predict_proba(X_test)
        #self.previsoes_real
        #self.previsoes_classificadores = predictions

        from sklearn.neural_network import MLPClassifier
        
        #clf.fit(modelHARD.previsoes_classificadores, y_val)
        #print("Deu certo")
        #print(clf.cv_results_)
        #print("Os resultados acima")
        #### Only stop true com proporção 0.005, 0.1, 0.2, ou false
        # -> deve ser o 'early_stopping' :['True','False']
        # -> deve ser 'validation_fraction': [0.005, 0.1, 0.2]
        ### numero de neuronios OK 'hidden_layer_sizes':[(50,), (100,), (200,), (300,)]
        ###taxa de aprendizado OK 'learning_rate':['constant', 'invscaling', 'adaptive']
        clf = MLPClassifier(hidden_layer_sizes=(10,), solver='adam', alpha=0.01, max_iter=10, random_state=1)
        #clf = MLPClassifier(hidden_layer_sizes=(200,), solver='adam', alpha=0.0001, random_state=1)
        #clf = MLPClassifier(solver='lbfgs', alpha=1e-5, hidden_layer_sizes=(5, 2), random_state=1)
        
        try:
            #clf.fit(modelHARD.previsoes_classificadores, modelHARD.previsoes_real)
            if int_k_inserir == 0:
                #TREINAR A REDE COM OS RECURSOS DE TREINO E VALIDAÇÃO
                '''
                if rede_grid:
                    clf.fit(modelHARD.previsoes_classificadores, y_val)
                    print("Resultado grid inicio")
                    print(clf.cv_results_)
                    print(clf.best_estimator_)
                    print(clf.best_params_)
                    print("Resultado grid fim")
                    rede_grid_estimador = clf.best_estimator_
                else:
                    clf = rede_grid_estimador
                    clf.fit(modelHARD.previsoes_classificadores, y_val)
                '''
                clf.fit(modelHARD.previsoes_classificadores, y_val)
                classificador_neural_softAUX = clf
            else:
                clf = classificador_neural_softAUX
            #print("Real")
            #print(y_test.iloc[self.dynamic_index_usage_neighbors_test])
            #print("Previsto")
            #print(clf.predict(self.previsoes_classificadores))
            #print("Probabilidade")
            #PREVENDO DADOS DE (TESTE OU VAL2)
            predicted_pesos  = clf.predict_proba(self.previsoes_classificadores)
            #predicted_pesos  = clf.predict_proba()
            #print("predicted_pesos")
            #print(predicted_pesos)
            #return predicted_pesos
            
        except:
            predicted_pesos = np.array([[0]])
        



        #print("Passou pelo erro")
        x, y, = predicted_pesos.shape
        ######if(len(np.unique(y, return_counts=False)) > y):
        #print("Ajuste rede 1")
        df_dados_x = pd.DataFrame()
        x = 1
        #for pool in self.estimators:
        for pool in pool_classifiers:#Corrrigir regerencia a pool_classifiers
           df_dados_x[str(x)] = pool.predict(X_val)
           x = x + 1
        '''
        print("###############################")
        print("Dados treino [0 1 0 1 0] em modelHARD  de X_val")
        print(modelHARD.previsoes_classificadores)
        print("###############################")
        print("Dados [0 1 0 1 0] de X_val em df_dados_x, usado para reinar a rede com y_val")
        print(df_dados_x)
        '''
        
        
        #print("Teste de Grid parametro")
        parameters = {'early_stopping':[False, True], 'validation_fraction': [0.005, 0.1, 0.2], 'max_iter':[500], 'hidden_layer_sizes':[(100,), (50,), (200,), (300,)], 'learning_rate':['constant', 'invscaling', 'adaptive'], 'solver':['sgd']}
        #print("Declarar  grid")
        clf = GridSearchCV(estimator = MLPClassifier(), scoring='f1_samples', param_grid = parameters, refit=True, cv = 5)
        #print("Treinar Grid")
 
        #print("self.previsoes_classificadores")
        #print(self.previsoes_classificadores)
        if int_k_inserir == 0:
            #TREINAR A REDE COM OS RECURSOS DE TREINO E VALIDAÇÃO
            if rede_grid:
                #print("Teste de Grid parametro")
                clf.fit(df_dados_x, y_val)
                print("Resultado grid Rede Soft")
                #print(clf.cv_results_)
                #print(clf.best_estimator_)
                print(clf.best_params_)
                #print("Resultado grid fim")
                rede_grid_estimador = clf.best_estimator_
                clf = rede_grid_estimador
            else:
                clf = rede_grid_estimador
                #clf.fit(modelHARD.previsoes_classificadores, y_val)
                clf.fit(df_dados_x, y_val)
            #clf = MLPClassifier(hidden_layer_sizes=(200,), solver='adam', alpha=0.0001, max_iter=1000, random_state=1)
            #clf.fit(df_dados_x, y_val)
            #clf.fit(modelHARD.previsoes_classificadores, y_val)
            classificador_neural_soft = clf 
        else:
            clf = classificador_neural_soft
        #print("o que classsificar")
        #print(self.previsoes_classificadores)
        predicted_pesos  = clf.predict_proba(self.previsoes_classificadores)
        #print("predicted_pesos")
        #print(predicted_pesos)
        #print (predicted_pesos)
        '''
        x, y, = predicted_pesos.shape
        if(len(self.classes_) > y):
            #print(y_val)
            predicted_pesos = self._mask_proba(pbb, slclsf)

        '''
        return predicted_pesos


    def _peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1(self,masked_proba):#Devolve as probabilidades mas remove dos classificadores não competentes
        predicted_proba = np.mean(masked_proba, axis=1)

        from sklearn.neighbors import KNeighborsClassifier
        neigh = KNeighborsClassifier(n_neighbors=len(y_val), n_jobs = 1)
        neigh.fit(X_val, y_val)


        row_kneighbors_validation = neigh.kneighbors(X=X_test, n_neighbors=None, return_distance=True)
        #print(X_val2)
        #print("Vizinhos do Teste (os vizinhos estão no validation 2)")
        #print(row_kneighbors_validation[0])
        #print(row_kneighbors_validation[1])
        #print("self.dynamic_index_usage_neighbors_test")
        #print(self.dynamic_index_usage_neighbors_test)

        #lista de vizinhos proximos das classes que serão classificadas, os valores estão na base de teste
        #row_unique_kneighbors_validation = np.unique(row_kneighbors_validation, return_counts=False)


        #print("predicted_proba resultado")
        #print(predicted_proba)
        dynamic_neighbors_distances_test = row_kneighbors_validation[0]
        dynamic_neighbors_index_train = row_kneighbors_validation[1]


        max_distance = max(np.unique(dynamic_neighbors_distances_test))
        #print("max_distance")
        #print(max_distance)
        #row_kneighbors_teste =
        # self.classes_.take(preds)
        #print("self.classes_")
        classes = self.classes_

        qt_classes = len(self.classes_)

        predicted_pesos = np.zeros(predicted_proba.shape)
        #print("self.dynamic_neighbors_index_train")
        #print(self.dynamic_neighbors_index_train)
        x, y = dynamic_neighbors_index_train.shape
        pesos_pela_distancia_para_classes = []
        for i in range(x):
            cont_class_neighbors_line = [0] * qt_classes
            distance_class_neighbors_line = [0] * qt_classes
            quant_div_distance_class_neighbors_line = [0] * qt_classes
            #print("cont_class_neighbors_line")
            #print(cont_class_neighbors_line)
            for j in range(y):
                for l in range(qt_classes):
                    if y_train.iloc[dynamic_neighbors_index_train[i,j]] == classes[l]:
                        #cont_class_neighbors_line[l] =  cont_class_neighbors_line[l] + 1
                        #print("3")
                        distance_class_neighbors_line[l] = distance_class_neighbors_line[l] + dynamic_neighbors_distances_test[i,j]
                        #print("4")
                        quant_div_distance_class_neighbors_line[l] = quant_div_distance_class_neighbors_line[l] + 1
                        #print("5")
            #print("JOTA SAI COM " + str(j))
            #print("distance_class_neighbors_line: " + str(distance_class_neighbors_line) )
            #print([(j+1)] * qt_classes)
            vizinhos_total_lista = ([(j+1)] * qt_classes)
            quant_div_distance_class_neighbors_line = [1 if value==0 else value for value in quant_div_distance_class_neighbors_line]# colocar o 1 para evitar divisão por 0
            mean_distance_class_neighbors_line = np.array(distance_class_neighbors_line) / np.array(quant_div_distance_class_neighbors_line)
            #print("mean_distance_class_neighbors_line: "+str(mean_distance_class_neighbors_line))
            #weight_distane_class_line = max_distance - mean_distance_class_neighbors_line
            #trocado por
            weight_distane_class_line = 1 / mean_distance_class_neighbors_line

            #print("weight_distane_class_line")
            #print(weight_distane_class_line)
            pesos_pela_distancia_para_classes.append(weight_distane_class_line)
            #print("-------------------------------------------------------")

            try:
                predicted_pesos[i] = predicted_proba[i] * (weight_distane_class_line * (1/np.sum(weight_distane_class_line)))
            except:
                pass

            #ou
            #predicted_pesos[i] = predicted_proba[i] * weight_distane_class_line
        #print("pesos_pela_distancia_para_classes")
        #print(pesos_pela_distancia_para_classes)
        #print("Resultado")
        #print(predicted_pesos)

        predicted_pesos = predicted_pesos / predicted_pesos.sum(axis=1)[:, None]
        #print("predicted_pesos resultado devolvido")
        #print(predicted_pesos)

        return predicted_pesos



    ############# SALVO #####
    def _sum_weight_0a1_votes_per_class(self, predictions):
        n_classes = len(self.classes_)
        #Votação ponderada por distancia em relação ao vizinho de teste mais distante (Geral dos teste)
        #Peso formatado aplicado na votação que o comitê deu
        #print("FUNÇÃO VOTO COM PESA POR CLASSE")
        #from sklearn.neighbors import KNeighborsClassifier
        #neigh = KNeighborsClassifier(n_neighbors=5, n_jobs = 1)
        #neigh.fit(X_val, y_val)
        #row_kneighbors_validation = neigh.kneighbors(X=X_test, n_neighbors=None, return_distance=False)
        #print("Vizinhos do Teste (os vizinhos estão no validation)")
        #print(row_kneighbors_validation)
        #lista de vizinhos proximos das classes que serão classificadas, os valores estão na base de teste
        #row_unique_kneighbors_validation = np.unique(row_kneighbors_validation, return_counts=False)
        #print("VALORES PARA CLASSIFICAR X")
        #MELHORAR AINDA ESTA CONSUMINDO MUITO
        #X_neighbors_val = X_val.iloc[row_unique_kneighbors_validation]
        #y_neighbors_val = y_val.iloc[row_unique_kneighbors_validation]
        from sklearn.neighbors import KNeighborsClassifier
        neigh = KNeighborsClassifier(n_neighbors=len(y_val), n_jobs = 1)
        neigh.fit(X_val, y_val)


        row_kneighbors_validation = neigh.kneighbors(X=X_test, n_neighbors=None, return_distance=True)
        #print(X_val2)
        #print("Vizinhos do Teste (os vizinhos estão no validation 2)")
        #print(row_kneighbors_validation[0])
        #print(row_kneighbors_validation[1])
        #print("self.dynamic_index_usage_neighbors_test")
        #print(self.dynamic_index_usage_neighbors_test)

        #lista de vizinhos proximos das classes que serão classificadas, os valores estão na base de teste
        #row_unique_kneighbors_validation = np.unique(row_kneighbors_validation, return_counts=False)


        #print("predicted_proba resultado")
        #print(predicted_proba)
        dynamic_neighbors_distances_test = row_kneighbors_validation[0]
        dynamic_neighbors_index_train = row_kneighbors_validation[1]

        max_distance = max(np.unique(dynamic_neighbors_distances_test))
        #print("max_distance")
        #print(max_distance)
        #row_kneighbors_teste =
        # self.classes_.take(preds)
        #print("self.classes_")
        classes = self.classes_
        votes = np.zeros((predictions.shape[0], n_classes), dtype=np.int64)
        #print("Votes instance sum votes (Cria amatriz zerada)")
        #print(votes)
        for label in range(n_classes):
            votes[:, label] = np.sum(predictions == label, axis=1)#Conta label nas coluna(eixo 1)
        #print("Votos majory resultado")
        #print(votes)
        #print("RESULTADO predicted_proba")
        #print(votes / votes.sum(axis=1)[:, None])
        #ponderacao = predicted_proba
        qt_classes = len(self.classes_)

        predicted_pesos = np.zeros(votes.shape)

        x, y = dynamic_neighbors_index_train.shape
        pesos_pela_distancia_para_classes = []
        for i in range(x):
            cont_class_neighbors_line = [0] * qt_classes
            distance_class_neighbors_line = [0] * qt_classes
            quant_div_distance_class_neighbors_line = [0] * qt_classes
            #print("cont_class_neighbors_line")
            #print(cont_class_neighbors_line)
            for j in range(y):
                for l in range(qt_classes):
                    #print("y_train[j] == classes[l]")
                    #print(y_train.iloc[self.dynamic_neighbors_index_train[i,j]])
                    #print(classes[l])
                    if y_train.iloc[dynamic_neighbors_index_train[i,j]] == classes[l]:
                        #cont_class_neighbors_line[l] =  cont_class_neighbors_line[l] + 1
                        distance_class_neighbors_line[l] = distance_class_neighbors_line[l] + dynamic_neighbors_distances_test[i,j]
                        quant_div_distance_class_neighbors_line[l] = quant_div_distance_class_neighbors_line[l] + 1
            #print("JOTA SAI COM " + str(j))
            #print("distance_class_neighbors_line: " + str(distance_class_neighbors_line) )
            #print([(j+1)] * qt_classes)
            vizinhos_total_lista = ([(j+1)] * qt_classes)
            quant_div_distance_class_neighbors_line = [1 if value==0 else value for value in quant_div_distance_class_neighbors_line]
            mean_distance_class_neighbors_line = np.array(distance_class_neighbors_line) / np.array(quant_div_distance_class_neighbors_line)
            #print("mean_distance_class_neighbors_line: "+str(mean_distance_class_neighbors_line))
            ##############weight_distane_class_line = max_distance - mean_distance_class_neighbors_line############
            weight_distane_class_line = 1 / mean_distance_class_neighbors_line

            #print("weight_distane_class_line")
            #print(weight_distane_class_line)
            pesos_pela_distancia_para_classes.append(weight_distane_class_line)
            #print("-------------------------------------------------------")

            #predicted_pesos[i] = votes[i] * weight_distane_class_line
            #OU
            try:
                predicted_pesos[i] = votes[i] * weight_distane_class_line * (1/np.sum(weight_distane_class_line))
            except:
                pass
            #predicted_pesos[i] = predicted_proba[i] * weight_distane_class_line * (1/np.sum(weight_distane_class_line))

        #print("pesos_pela_distancia_para_classes")
        #print(pesos_pela_distancia_para_classes)
        #print("Resultado")
        #print(predicted_pesos)

        predicted_pesos = predicted_pesos / predicted_pesos.sum(axis=1)[:, None]
        #print("predicted_pesos")
        #print(predicted_pesos)
        return predicted_pesos


    # Borda por meio de Votação



    def naive_bayes_combination(self, predictions, n_classes):
        #print("FUNÇÃO naive_bayes_combination")
        from sklearn.metrics import confusion_matrix
        m_confusoes = []
        for m in range(len(pool_classifiers)):#CORRIGIR
            m_confusoes.append(confusion_matrix(y_val, pool_classifiers[m].predict(X_val),labels=self.classes_))
        cont_elementos = [0] * len(self.classes_)
        for i in range(len(self.classes_)):
            for v in range(len(y_val)):
                if(y_val.iloc[v] == self.classes_[i]):
                    cont_elementos[i] = cont_elementos[i] + 1
        #print(predictions)
        xxx, yyy = predictions.shape
        predicted_proba_resultado = []
        for x in range(xxx):
            resultado = [1] * len(self.classes_)
            for y in range(yyy):
                if (predictions[x,y] >= 0):
                    matriz_c = m_confusoes[y]
                    #print(matriz_c)
                    #print(matriz_c[:,predictions[x,y]])
                    resultado = resultado * matriz_c[:,predictions[x,y]]
            class_resultado = resultado / cont_elementos
            if(class_resultado.sum() >0 ):#Erro de quando for 0
                predicted_pesos = class_resultado / class_resultado.sum()
            else:
                predicted_pesos = [(1/len(self.classes_))] * len(self.classes_)
            predicted_proba_resultado.append([predicted_pesos])
        #return predicted_proba_resultado
        return np.concatenate(predicted_proba_resultado)

def _max_proba(probabilities):#Devolve as probabilidades mas remove dos classificadores não competentes
    predicted_proba = np.amax(probabilities, axis=1)
    #print("predicted_proba in function _max_proba using np.amax")
    #print(predicted_proba)

    normalizar_linha = 1 / predicted_proba.sum(axis=1)[:, None]
    #print("normalizar_linha in function _max_proba using 1 / predicted_proba.sum(axis=1)[:, None]")
    #print(normalizar_linha)
    predicted_proba = predicted_proba * normalizar_linha

    #print("predicted_proba resultado")
    #print(predicted_proba)

    return predicted_proba


def _minimun_proba(probabilities):#Devolve as probabilidades mas remove dos classificadores não competentes
    predicted_proba = np.amin(probabilities, axis=1)
    #print("predicted_proba in function _minimun_proba using np.amax")
    #print(predicted_proba)

    normalizar_linha = 1 / predicted_proba.sum(axis=1)[:, None]
    #print("normalizar_linha in function _minimun_proba using 1 / predicted_proba.sum(axis=1)[:, None]")
    #print(normalizar_linha)
    predicted_proba = predicted_proba * normalizar_linha

    #print("predicted_proba resultado")
    #print(predicted_proba)

    return predicted_proba

def borda_class(masked_proba):#Devolve as probabilidades mas remove dos classificadores não competentes
    #BORDA
    x, y, z = masked_proba.shape
    #Grupos(exemplos), lindas por grupos(classificadores), colunas(classes)
    #print("x y z")
    #print(x,y,z)
    for i in range(x):
        for j in range(y):
            for k in range(z):
                max_linha = np.amax(masked_proba[i,j])
                for sub in range(z):
                    if masked_proba[i,j,sub] == max_linha :
                        masked_proba[i,j,sub] = (z - k -1) * -1
                        break
    #print("masked_proba no final")
    #print(masked_proba)
    #print("masked_proba no final")
    #print(masked_proba * -1)
    masked_proba = masked_proba * -1
    #print("Outra")
    #print(np.sum(masked_proba,axis=1))
    masked_proba = np.sum(masked_proba,axis=1)
    #print(" masked_proba.sum(axis=1)[:, None]")
    #print( masked_proba.sum(axis=1)[:, None])
    #print(" masked_proba.sum(axis=1)")
    #print( masked_proba.sum(axis=1))
    predicted_pesos = masked_proba / masked_proba.sum(axis=1)[:, None]
    #print("predicted_pesos")
    #print(predicted_pesos)
    return predicted_pesos

def geometric_mean(probabilities):#Devolve as probabilidades mas remove dos classificadores não competentes
    predicted_proba = np.prod(probabilities, axis=1)
    #print("predicted_proba in function media_geometrica using np.prod(masked_proba, axis=1)")
    #print(predicted_proba)
    normalizar_linha = 1 / predicted_proba.sum(axis=1)[:, None]
    #print("normalizar_linha in function media_geometrica using 1 / predicted_proba.sum(axis=1)[:, None]")
    #print(normalizar_linha)
    predicted_proba = predicted_proba * normalizar_linha
    #print("predicted_proba resultado")
    #print(predicted_proba)
    return predicted_proba

import numpy as np
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import RandomForestClassifier
#clf1 = LogisticRegression(multi_class='multinomial', random_state=1)
#clf2 = RandomForestClassifier(n_estimators=50, random_state=1)
#clf3 = GaussianNB()
#X = np.array([[-1, -1], [-2, -1], [-3, -2], [1, 1], [2, 1], [3, 2]])
#y = np.array([1, 1, 1, 2, 2, 2])

'''
#Maternal Health Risk         #https://archive.ics.uci.edu/ml/datasets/Maternal+Health+Risk+Data+Set
qualidade = pd.read_csv("bases/Maternal_Health_Risk.csv")
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('RiskLevel',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('RiskLevel',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('RiskLevel',axis=1).columns )
X = df_feat
y = qualidade['RiskLevel']
'''
'''
#PROGNOSTIC
qualidade = pd.read_csv("bases/wpbc.data", header=None, prefix='col_')
qualidade = qualidade.drop('col_0',axis=1)
#print(qualidade)

scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('col_1',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('col_1',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('col_1',axis=1).columns )
X = df_feat
y = qualidade['col_1']
'''

'''
#ZOO
qualidade = pd.read_csv("bases/zoo.data")
qualidade = qualidade.drop('animal',axis=1)
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('type',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('type',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('type',axis=1).columns )
X = df_feat
y = qualidade['type']
'''

'''
#BUPA
qualidade = pd.read_csv("bases/bupa.data")
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('class',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('class',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('class',axis=1).columns )
X = df_feat
y = qualidade['class']
'''

'''
#Steel Plates Faults   https://archive.ics.uci.edu/ml/datasets/Steel+Plates+Faults NÂO USADO
qualidade = pd.read_csv("bases/faults.csv", header=None, prefix='col_')
#qualidade = qualidade.drop('col_69',axis=1)
#print(qualidade)
#print("Nulos")
#print(qualidade.isnull().sum().sort_values(ascending=False))
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('col_27',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('col_27',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('col_27',axis=1).columns )
X = df_feat
y = qualidade['col_27']
'''

'''
#BREAST 682
qualidade = pd.read_excel("bases/BreastTissue.xls",sheet_name="Data")
qualidade = qualidade.drop('Case #',axis=1)
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('Class',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('Class',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('Class',axis=1).columns )
X = df_feat
y = qualidade['Class']
'''

'''
#LIBRAS
qualidade = pd.read_csv("bases/movement_libras.data")
#print(qualidade)

scaler = StandardScaler() # inicializar
scaler.fit(qualidade.iloc[:,0:-1])#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.iloc[:,0:-1])#normaliza por media edesvio padrão
X = pd.DataFrame(scaled_features)
#print(X)
#X = qualidade.iloc[:,0:-1]
y = qualidade.iloc[:,-1]
'''

'''
#Congressional Voting Records 772 TESTAR ESSE AQUI
qualidade = pd.read_csv("bases/congressional_voting_records.data")

scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('class',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('class',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('class',axis=1).columns )
X = df_feat
y = qualidade['class']
'''

'''
#Algerian Forest Fires 
qualidade = pd.read_csv("bases/Algerian_forest_fires.csv")
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('Classes',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('Classes',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('Classes',axis=1).columns )
X = df_feat
y = qualidade['Classes']
'''
'''
#Maternal Health Risk         #https://archive.ics.uci.edu/ml/datasets/Maternal+Health+Risk+Data+Set
qualidade = pd.read_csv("bases/Maternal_Health_Risk.csv")
#qualidade = qualidade.drop('col_69',axis=1) 
#print(qualidade)
#print("Nulos")
#print(qualidade.isnull().sum().sort_values(ascending=False))

            
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('RiskLevel',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('RiskLevel',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('RiskLevel',axis=1).columns )
X = df_feat
y = qualidade['RiskLevel']
#print(X)
#print(y)
'''
'''
#PROGNOSTIC
qualidade = pd.read_csv("bases/wpbc.data", header=None, prefix='col_')
qualidade = qualidade.drop('col_0',axis=1)
#print(qualidade)

scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('col_1',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('col_1',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('col_1',axis=1).columns )
X = df_feat
y = qualidade['col_1']
'''

'''
#WINE
qualidade = pd.read_csv("bases/wine.data")
#qualidade = qualidade.drop('id',axis=1) 
#print(qualidade)
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('class',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('class',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('class',axis=1).columns )
X = df_feat
y = qualidade['class']
'''

#IONOSPHERE 232
qualidade = pd.read_csv("bases/ionosphere.data", header=None, prefix='col_')
#print(qualidade)
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('col_34',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('col_34',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('col_34',axis=1).columns )
X = df_feat
y = qualidade['col_34']


'''
CAR
qualidade = pd.read_csv("bases/car2.data")
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('class',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('class',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.columns[:-1])
X = df_feat
y = qualidade['class']
'''
'''
#Dermatology
qualidade = pd.read_csv("bases/dermatology.data", header=None, prefix='col_')
median_col_33 = qualidade.col_33.median()
#print(median_col_33)
qualidade.fillna(median_col_33, inplace=True)
#print(qualidade.isnull().sum().sort_values(ascending=False))
           
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('col_34',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('col_34',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('col_34',axis=1).columns )
X = df_feat
y = qualidade['col_34']
'''

'''
#Risk Factors Cervical Cancer  https://archive.ics.uci.edu/ml/datasets/Cervical+cancer+%28Risk+Factors%29 862
qualidade = pd.read_csv("bases/risk_factors_cervical_cancer.csv")
qualidade = qualidade.drop('Hinselmann',axis=1)
qualidade = qualidade.drop('Schiller',axis=1)
qualidade = qualidade.drop('Citology',axis=1)
qualidade = qualidade.drop('STDs: Time since last diagnosis',axis=1)
qualidade = qualidade.drop('STDs: Time since first diagnosis',axis=1)
qualidade = qualidade.drop('STDs:AIDS',axis=1)
qualidade = qualidade.drop('STDs:cervical condylomatosis',axis=1)
#print(qualidade)
#print("Nulos")
#print(qualidade.isnull().sum().sort_values(ascending=False))
varlor = pd.DataFrame(qualidade.isnull().sum().sort_values(ascending=False))
for col_nula in range(len(varlor)):
    if( varlor.iloc[col_nula,0] > 0):
        #print(str(varlor.iloc[col_nula].name))
        median_col = qualidade[str(varlor.iloc[col_nula].name)].median()
        desvio_col = qualidade[str(varlor.iloc[col_nula].name)].std()
        #print(str(median_col)+" - " + str(desvio_col))
        qualidade[str(varlor.iloc[col_nula].name)].fillna(median_col, inplace=True)
scaler = StandardScaler() # inicializar
scaler.fit(qualidade.drop('Biopsy',axis=1))#treinar o modelo de normalização para padronizar,
scaled_features = scaler.transform(qualidade.drop('Biopsy',axis=1))#normaliza por media edesvio padrão
df_feat = pd.DataFrame(scaled_features,columns=qualidade.drop('Biopsy',axis=1).columns )
X = df_feat
y = qualidade['Biopsy']
'''
            
v_total_k = 0

#v_total_teste = [5]
v_total_teste = [5,10,15,20,25,30,60]
#v_total_teste = [5,10,15,20,25,30,60]
for vizinhos_no_teste in v_total_teste:
    rede_grid = True
    print("")
    print("------------------- K do POOL  "+ str(vizinhos_no_teste) +" de "+str(v_total_teste)  +" --------------------------------------------------------------------")

    #for melhor_k in [3,7,11]:
    for melhor_k in [3]:
        # [Tres vetores]  Matriz em cada porcentagem, para guardas as listas dos tres valores
        porcentagem_escolheu_MAX= [[],[],[]]
        porcentagem_escolheu_SOFT= [[],[],[]]
        porcentagem_escolheu_HARD= [[],[],[]]
        porcentagem_escolheu_MIN= [[],[],[]]
        porcentagem_escolheu_G_MEAN= [[],[],[]]
        porcentagem_escolheu_sum_weight = [[],[],[]]
        porcentagem_escolheu_rede_neural= [[],[],[]]
        porcentagem_escolheu_rede_neural_soft = [[],[],[]]
        porcentagem_escolheu_rede_neural_soft_div = [[],[],[]]
        porcentagem_escolheu_borda = [[],[],[]]
        porcentagem_escolheu_naive_bayes = [[],[],[]]
        porcentagem_escolheu_peso_ponderado_comite = [[],[],[]]
        porcentagem_escolha_compartilhada = [[],[],[]]
        #porcentagem_escolheu_MAX, porcentagem_escolheu_SOFT, porcentagem_escolheu_HARD, porcentagem_escolheu_MIN, porcentagem_escolheu_G_MEAN, porcentagem_escolheu_sum_weight , porcentagem_escolheu_rede_neural, porcentagem_escolheu_rede_neural_soft , porcentagem_escolheu_rede_neural_soft_div , porcentagem_escolheu_borda , porcentagem_escolheu_naive_bayes , porcentagem_escolheu_peso_ponderado_comite = [],[],[],[], [],[],[],[],[], [],[],[]
        #porcentagem_escolheu_MAX, porcentagem_escolheu_SOFT, porcentagem_escolheu_HARD, porcentagem_escolheu_MIN, porcentagem_escolheu_G_MEAN, porcentagem_escolheu_sum_weight, porcentagem_escolheu_rede_neural, porcentagem_escolheu_rede_neural_soft, porcentagem_escolheu_rede_neural_soft_div, porcentagem_escolheu_borda, porcentagem_escolheu_naive_bayes, porcentagem_escolheu_peso_ponderado_comite
        geral_empate_classificadores_no_inicio = [[],[],[]]
        geral_empate_classificadores_no_final = [[],[],[]]
        resultados_no_melhor_caso = []
        resultados_hard = []
        resultados_soft = []
        resultados_max = []
        resultados_min = []
        resultados_geometric_mean = []
        quantidade_exemplos_divergencia = []
        quantidade_classificadores_selecionados = []
        pesos_dos_classificadores_selecionados = []
        resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1 = []
        resultados_sum_weight_0a1_votes_per_class = []
        resultados_dynamic_metric_fusionk3 = [[],[],[]]
        resultados_escolheu_rede_neural  = []
        resultados_escolheu_rede_neural_soft  = []
        resultados_escolheu_rede_neural_soft_div  = []
        resultados_escolheu_borda  = []
        resultados_escolheu_naive_bayes  = []
        resultados_maximo_na_combinacao = []





        #int_k_inserir = 1

        #print("------------------------------ INICIO K = "+ str(melhor_k)+" ------------------------------")
        k_dynamic_combination = melhor_k
        total_execucoes = 30#VALIDAÇÕES, EXECUCOES
        for iii in range(total_execucoes):
            #v_total_k = melhor_k#Da fusão dinamica
            int_k_inserir = 0#Inserir na matriz com 0
            #print("")
            if(iii==0) or (iii==1) or (iii==2) or (iii==4) or (iii==9) or (iii==19) or (iii==29):
                print("----------------------- Teste " + str(iii+1)+ " De "+str(total_execucoes)+" ------------------------------------------")


            #0.25
            X_train, X_test, y_train, y_test = train_test_split(X, #train.drop('Survived',axis=1) remove a coluna e a retorna mas não altera na variavel
                                                                y, test_size=0.50,#y, test_size=0.5,
                                                                random_state=(iii+vizinhos_no_teste+46))#Dividir os dados em treino e teste

            #0.333333333
            X_test, X_val, y_test, y_val = train_test_split(X_test, #train.drop('Survived',axis=1) remove a coluna e a retorna mas não altera na variavel
                                                            y_test, test_size=0.333333,# y_test, test_size=0.33333,
                                                            random_state=(iii+vizinhos_no_teste+34534))#Dividir os dados em treino e  validaçã

            X_test, X_val2, y_test, y_val2 = train_test_split(X_test, #train.drop('Survived',axis=1) remove a coluna e a retorna mas não altera na variavel
                                                              y_test, test_size=0.5,#y_test, test_size=0.5,
                                                              random_state=(iii+vizinhos_no_teste+65464))#Dividir os dados em treino e  validaçã



            inc = 0
            #while( len(np.unique(y_train, return_counts=False)) < len(np.unique(y, return_counts=False)) or len(np.unique(y_val, return_counts=False)) < len(np.unique(y, return_counts=False)) or len(np.unique(y_val2, return_counts=False)) < len(np.unique(y, return_counts=False))or len(np.unique(y_test, return_counts=False)) < len(np.unique(y, return_counts=False))):
            while( len(np.unique(y_train, return_counts=False)) < len(np.unique(y, return_counts=False)) or len(np.unique(y_val, return_counts=False)) < len(np.unique(y, return_counts=False))):
                #if(inc % 70 == 0):
                #    print("Reajustado o valor de Y train")
                inc = inc + 7
                #0.25
                X_train, X_test, y_train, y_test = train_test_split(X, #train.drop('Survived',axis=1) remove a coluna e a retorna mas não altera na variavel
                                                                    y, test_size=0.5,
                                                                    random_state=(inc+vizinhos_no_teste+iii+5675))#Dividir os dados em treino e teste
                #0.333333333
                X_test, X_val, y_test, y_val = train_test_split(X_test, #train.drop('Survived',axis=1) remove a coluna e a retorna mas não altera na variavel
                                                                  y_test, test_size=0.33333,
                                                                  random_state=(inc+vizinhos_no_teste+iii+7675))#Dividir os dados em treino e  validaçã

                X_test, X_val2, y_test, y_val2 = train_test_split(X_test, #train.drop('Survived',axis=1) remove a coluna e a retorna mas não altera na variavel
                                                                  y_test, test_size=0.50,
                                                                  random_state=(inc+ vizinhos_no_teste+iii+345))#Dividir os dados em treino e  validaçã





            
            parameters = {'n_neighbors':[5], 'weights':['uniform', 'distance'], 'algorithm': ['ball_tree','kd_tree','brute'], 'leaf_size':[10,15,30,45,60,120], 'p':[2,1,3]}
            #print("Declarar  grid")
            grid_KNN = GridSearchCV(estimator = KNeighborsClassifier(),scoring='f1_samples', param_grid = parameters, refit=True, cv = 10)
            if knn_is_grid:
                #print("Teste de Grid parametros no KNN")
                grid_KNN.fit(X_val2, y_val2)
                #print("Resultado grid KNN vvv")
                #print(clf.cv_results_)
                #print(grid_KNN.best_estimator_)
                if(iii==0) or (iii==2) or (iii==4) or (iii==9) or (iii==19) or (iii==29):
                    print(grid_KNN.best_params_)
                #print("Resultado grid KNN ^^^^")
                knn_best_clf_in_grid = grid_KNN.best_estimator_
                #knn_is_grid = False
        
        

            if(iii == 0):
                print("")
                print("Treino - Teste - Validação 1 - Validação 2 - ")
                print(str(len(y_train))+ "    -   "+str(len(y_test))+"    -    "+str(len(y_val)) +"    -    "+str(len(y_val2)) )
                print("Número de Exemplos: " +str(len(X)))
                print("Número de Colunas: " +str(len(X.columns)))
                print("Número de classes: " + str(len(np.unique(y, return_counts=False))))
                print("")

            if(iii==0) or (iii==2) or (iii==4) or (iii==9) or (iii==19) or (iii==29):
                print("Classes - treino: " + str(len(np.unique(y_train, return_counts=False))) + " - teste: " + str(len(np.unique(y_test, return_counts=False)))+ " - val: " + str(len(np.unique(y_val, return_counts=False)))+ " - val_2: " + str(len(np.unique(y_val2, return_counts=False))))
                print("")

            '''
            #eclf1 = EnsembleClassifier(estimators=[('lr', clf1), ('rf', clf2), ('gnb', clf3)], voting='hard')
            eclf1 = EnsembleClassifier(estimators=[('lr', clf1), ('rf', clf2), ('gnb', clf3)], voting='hard', voting_type="vm")
            print("Voto Majoritario")
            eclf1 = eclf1.fit(X_train, y_train)
            print(eclf1.score(X_test, y_test))


            eclf1 = EnsembleClassifier(estimators=[('lr', clf1), ('rf', clf2), ('gnb', clf3)], voting='soft', voting_type="soft" )
            print("Média")
            eclf1 = eclf1.fit(X_train, y_train)
            print(eclf1.score(X_test, y_test))

            print("Máximo")
            eclf1 = EnsembleClassifier(estimators=[('lr', clf1), ('rf', clf2), ('gnb', clf3)], voting='soft', voting_type="max" )
            eclf1 = eclf1.fit(X_train, y_train)
            print(eclf1.score(X_test, y_test))

            print("Mínimo")
            eclf1 = EnsembleClassifier(estimators=[('lr', clf1), ('rf', clf2), ('gnb', clf3)], voting='soft', voting_type="min" )
            eclf1 = eclf1.fit(X_train, y_train)
            print(eclf1.score(X_test, y_test))


            eclf1 = EnsembleClassifier(estimators=[('lr', clf1), ('rf', clf2), ('gnb', clf3)], voting='soft', voting_type="rna" )
            print("Rede Neural Artificial")
            eclf1 = eclf1.fit(X_train, y_train)
            print(eclf1.score(X_test, y_test))
            '''
            #print("TESTE DE  FUSÃO DINAMICA")

            from sklearn.tree import DecisionTreeClassifier
            from sklearn.ensemble import BaggingClassifier
            rng = np.random.RandomState(42)

            modelp = DecisionTreeClassifier(random_state=rng, max_depth=10)
            pool_classifiers = BaggingClassifier(modelp, n_estimators=vizinhos_no_teste)
            pool_classifiers.fit(X_train, y_train)

            pool_esti = []
            for c in pool_classifiers:
                #print("Inicio")
                pool_esti.append((str(c),c))
                #print(c.predict(X_test))
                #print(c.predict_proba(X_test))
                #print("fim")

            '''
            print("#####################")
            modelSOFT = EnsembleClassifier(estimators=pool_esti, voting='soft', voting_type="soft")
            modelSOFT.fit(X_train,y_train )
            so = VotingClassifier(estimators=pool_esti, voting='soft')
            so.fit(X_train,y_train)
            print("Soft")
            print(modelSOFT.score(X_test, y_test))
            print(so.score(X_test, y_test))
            print(modelSOFT.predict(X_test))
            print(modelSOFT.predict_proba(X_test))
            print("Hard")
            modelHARD = EnsembleClassifier(estimators=pool_esti, voting='hard', voting_type="soft")
            ha = VotingClassifier(estimators=pool_esti, voting='hard')
            ha.fit(X_train,y_train )
            modelHARD.fit(X_train,y_train )
            print(modelHARD.score(X_test, y_test))
            print(ha.score(X_test, y_test))
            print(modelHARD.predict(X_test))
            print(modelHARD.predict_proba(X_test))
            print("#######################")
            '''
            
            #eclf1 = EnsembleClassifier(estimators=[('lr', clf1), ('rf', clf2), ('gnb', clf3)], voting='soft', voting_type="dynamic_fusion" )
            eclf1 = EnsembleClassifier(estimators=pool_esti, voting='soft', voting_type="dynamic_fusion" )
            #print("Dynamic fusion")
            eclf1 = eclf1.fit(X_train, y_train)
            #print(eclf1.score(X_test, y_test))



            #print(eclf1.predict(X_test))
            #print("Previsão da fusão dinamica")

            #print("Real")
            #print(y_test)
            if(iii==0) or (iii==2) or (iii==4) or (iii==9) or (iii==19) or (iii==29):
                print("------------------------------ INICIO K = 3 ------------------------------")

            v_total_k = 3
            o_resultado = eclf1.score(X_test, y_test)
            #print(o_resultado)
            resultados_dynamic_metric_fusionk3[int_k_inserir].append(o_resultado)
            #resultados_dynamic_metric_fusionk3.append(0)
            #print("Resultado dynamic_fusion")
            int_k_inserir = 1#Inserir na matriz com 1
            if(iii==0) or (iii==2) or (iii==4) or (iii==9) or (iii==19) or (iii==29):
                print("------------------------------ INICIO K = 7 ------------------------------")

            
            v_total_k = 7
            o_resultado = eclf1.score(X_test, y_test)
            #print(o_resultado)
            resultados_dynamic_metric_fusionk3[int_k_inserir].append(o_resultado)
            if(iii==0) or (iii==2) or (iii==4) or (iii==9) or (iii==19) or (iii==29):

                print("------------------------------ INICIO K = 11 ------------------------------")

            int_k_inserir = 2#Inserir na matriz com 2
            
            

            v_total_k = 11
            o_resultado = eclf1.score(X_test, y_test)
            #print(o_resultado)
            resultados_dynamic_metric_fusionk3[int_k_inserir].append(o_resultado)


        print("")
        print("")

        col_valor = get_column_letter(coluna_valor)
        col_desvio = get_column_letter(coluna_desvio_padrao)
        lin_inserir = linha_lida

        print("Voto Majoritário" + str(resultados_hard))
        #print(resultados_hard)
        print("Média: " + str(np.mean(resultados_hard)*100))
        print("Desvio padrão: " + str(np.std(resultados_hard)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_hard)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_hard)*100
        lin_inserir = lin_inserir + 1

        print("Média")
        #print(resultados_soft)
        print("Média: " + str(np.mean(resultados_soft)*100))
        print("Desvio padrão: " + str(np.std(resultados_soft)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_soft)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_soft)*100
        lin_inserir = lin_inserir + 1

        print("Máximo")
        #print(resultados_max)
        print("Média: " + str(np.mean(resultados_max)*100))
        print("Desvio padrão: " + str(np.std(resultados_max)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_max)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_max)*100
        lin_inserir = lin_inserir + 1

        print("Mínimo")
        #print(resultados_min)
        print("Média: " + str(np.mean(resultados_min)*100))
        print("Desvio padrão: " + str(np.std(resultados_min)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_min)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_min)*100
        lin_inserir = lin_inserir + 1

        print("Média geométrica")
        #print(resultados_geometric_mean)
        print("Média: " + str(np.mean(resultados_geometric_mean)*100))
        print("Desvio padrão: " + str(np.std(resultados_geometric_mean)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_geometric_mean)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_geometric_mean)*100
        lin_inserir = lin_inserir + 1

        #print("Peso_ponderado_classe_cada_amostra_sem_ajustes")
        #print(resultados_peso_ponderado_classe_cada_amostra_sem_ajustes)
        #print("Média: " + str(np.mean(resultados_peso_ponderado_classe_cada_amostra_sem_ajustes)*100))
        #print("Desvio padrão: " + str(np.std(resultados_peso_ponderado_classe_cada_amostra_sem_ajustes)*100))
        #print("")

        #print("Peso_ponderado_classe_cada_amostra_ajustado_0a1")
        #print(resultados_peso_ponderado_classe_cada_amostra_ajustado_0a1)
        #print("Média: " + str(np.mean(resultados_peso_ponderado_classe_cada_amostra_ajustado_0a1)*100))
        #print("Desvio padrão: " + str(np.std(resultados_peso_ponderado_classe_cada_amostra_ajustado_0a1)*100))
        #print("")

        #print("Peso_ponderado_comite_classe_distancia_maxima_teste")
        #print(resultados_peso_ponderado_comite_classe_distancia_maxima_teste)
        #print("Média: " + str(np.mean(resultados_peso_ponderado_comite_classe_distancia_maxima_teste)*100))
        #print("Desvio padrão: " + str(np.std(resultados_peso_ponderado_comite_classe_distancia_maxima_teste)*100))
        #print("")

        print("Peso ponderado no comitê")
        #print(resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1)
        print("Média: " + str(np.mean(resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1)*100))
        print("Desvio padrão: " + str(np.std(resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_peso_ponderado_comite_classe_distancia_maxima_teste_ajustado_0a1)*100
        lin_inserir = lin_inserir + 1


        #print("Sum_weight_votes_per_class")
        #print(resultados_sum_weight_votes_per_class)
        #print("Média: " + str(np.mean(resultados_sum_weight_votes_per_class)*100))
        #print("Desvio padrão: " + str(np.std(resultados_sum_weight_votes_per_class)*100))
        #print("")

        print("Votação ponderada no comitê")
        #print(resultados_sum_weight_0a1_votes_per_class)
        print("Média: " + str(np.mean(resultados_sum_weight_0a1_votes_per_class)*100))
        print("Desvio padrão: " + str(np.std(resultados_sum_weight_0a1_votes_per_class)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_sum_weight_0a1_votes_per_class)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_sum_weight_0a1_votes_per_class)*100
        lin_inserir = lin_inserir + 1

        #print("Sum_weight_line_votes_per_class")
        #print(resultados_sum_weight_line_votes_per_class)
        #print("Média: " + str(np.mean(resultados_sum_weight_line_votes_per_class)*100))
        #print("Desvio padrão: " + str(np.std(resultados_sum_weight_line_votes_per_class)*100))
        #print("")

        #print("Sum_weight_0a1_line_votes_per_class")
        #print(resultados_sum_weight_0a1_line_votes_per_class)
        #print("Média: " + str(np.mean(resultados_sum_weight_0a1_line_votes_per_class)*100))
        #print("Desvio padrão: " + str(np.std(resultados_sum_weight_0a1_line_votes_per_class)*100))
        #print("")



        #############
        print("Rede neural HARD")
        #print(resultados_escolheu_rede_neural)
        print("Média: " + str(np.mean(resultados_escolheu_rede_neural)*100))
        print("Desvio padrão: " + str(np.std(resultados_escolheu_rede_neural)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_escolheu_rede_neural)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_escolheu_rede_neural)*100
        lin_inserir = lin_inserir + 1

        print("Rede neural SOFT")
        #print(resultados_escolheu_rede_neural_soft)
        print("Média: " + str(np.mean(resultados_escolheu_rede_neural_soft)*100))
        print("Desvio padrão: " + str(np.std(resultados_escolheu_rede_neural_soft)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_escolheu_rede_neural_soft)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_escolheu_rede_neural_soft)*100
        lin_inserir = lin_inserir + 1

        #print("Rede neural SOFT CLASS")
        #print(resultados_escolheu_rede_neural_soft_div)
        #print("Média: " + str(np.mean(resultados_escolheu_rede_neural_soft_div)*100))
        #print("Desvio padrão: " + str(np.std(resultados_escolheu_rede_neural_soft_div)*100))
        #print("")
        #ws[col_valor + str(lin_inserir)] = np.mean(resultados_escolheu_rede_neural_soft_div)*100
        #ws[col_desvio + str(lin_inserir)] = np.std(resultados_escolheu_rede_neural_soft_div)*100
        lin_inserir = lin_inserir + 1

        print("Borda")
        #print(resultados_escolheu_borda)
        print("Média: " + str(np.mean(resultados_escolheu_borda)*100))
        print("Desvio padrão: " + str(np.std(resultados_escolheu_borda)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_escolheu_borda)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_escolheu_borda)*100
        lin_inserir = lin_inserir + 1

        print("Naive bayes")
        #print(resultados_escolheu_naive_bayes)
        print("Média: " + str(np.mean(resultados_escolheu_naive_bayes)*100))
        print("Desvio padrão: " + str(np.std(resultados_escolheu_naive_bayes)*100))
        print("")
        ws[col_valor + str(lin_inserir)] = np.mean(resultados_escolheu_naive_bayes)*100
        ws[col_desvio + str(lin_inserir)] = np.std(resultados_escolheu_naive_bayes)*100
        lin_inserir = lin_inserir + 1


        #print("resultados_dynamic_metric_fusion k-1")
        #print(resultados_dynamic_metric_fusionk1)
        #print("Média: " + str(np.mean(resultados_dynamic_metric_fusionk1)*100))
        #print("Desvio padrão: " + str(np.std(resultados_dynamic_metric_fusionk1)*100))
        #print("")

        ################### TRES VEZES
        for ii in range(3):
            print("Fusão Dinâmica K "+str(ii+1))
            
            #print(resultados_dynamic_metric_fusionk3)
            print("Média: " + str(np.mean(resultados_dynamic_metric_fusionk3[ii])*100))
            print("Desvio padrão: " + str(np.std(resultados_dynamic_metric_fusionk3[ii])*100))
            print(resultados_dynamic_metric_fusionk3[ii])
            print("")
            print("")
            ws[col_valor + str(lin_inserir)] = np.mean(resultados_dynamic_metric_fusionk3[ii])*100
            ws[col_desvio + str(lin_inserir)] = np.std(resultados_dynamic_metric_fusionk3[ii])*100
            #lin_inserir = lin_inserir + 1
            coluna_valor = coluna_valor + 2
            coluna_desvio_padrao = coluna_desvio_padrao +2
            col_valor = get_column_letter(coluna_valor)
            col_desvio = get_column_letter(coluna_desvio_padrao)


        coluna_valor = coluna_valor - 6
        coluna_desvio_padrao = coluna_desvio_padrao - 6
        col_valor = get_column_letter(coluna_valor)
        col_desvio = get_column_letter(coluna_desvio_padrao)
        #col_desvio = col_desvio - 6




        lin_inserir = lin_inserir + 2





        '''
        print("resultados_dynamic_metric_fusionk5")
        #print(resultados_dynamic_metric_fusionk5)
        print("Média: " + str(np.mean(resultados_dynamic_metric_fusionk5)*100))
        print("Desvio padrão: " + str(np.std(resultados_dynamic_metric_fusionk5)*100))
        print("")

        print("resultados_dynamic_metric_fusion k-7")
        #print(resultados_dynamic_metric_fusionk7)
        print("Média: " + str(np.mean(resultados_dynamic_metric_fusionk7)*100))
        print("Desvio padrão: " + str(np.std(resultados_dynamic_metric_fusionk7)*100))
        print("")

        print("resultados_dynamic_metric_fusionk13")
        #print(resultados_dynamic_metric_fusionk13)
        print("Média: " + str(np.mean(resultados_dynamic_metric_fusionk13)*100))
        print("Desvio padrão: " + str(np.std(resultados_dynamic_metric_fusionk13)*100))
        print("")


        print("resultados_dynamic_metric_fusion k-21")
        #print(resultados_dynamic_metric_fusionk21)
        print("Média: " + str(np.mean(resultados_dynamic_metric_fusionk21)*100))
        print("Desvio padrão: " + str(np.std(resultados_dynamic_metric_fusionk21)*100))
        print("")
        '''


        ############################### TRES VEZES
        for ii in range(3):
            print("###########  "+ str(ii)+"  ################")
            
            print("Média de porcentagem que escolheu Média: " + str(np.mean(porcentagem_escolheu_SOFT[ii])))
            #print("Desvio padrão de porcentagem que escolheu Média: " + str(np.std(porcentagem_escolheu_SOFT)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_SOFT[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_SOFT)
            lin_inserir = lin_inserir + 1
            #print("")

            print("Média de porcentagem que escolheu Voto Majoritário: " + str(np.mean(porcentagem_escolheu_HARD[ii])))
            #print("Desvio padrão de porcentagem que escolheu Voto Majoritário: " + str(np.std(porcentagem_escolheu_HARD)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_HARD[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_HARD)
            lin_inserir = lin_inserir + 1
            #print("")

            print("Média de porcentagem que escolheu Média geométrica: " + str(np.mean(porcentagem_escolheu_G_MEAN[ii])))
            #print("Desvio padrão de porcentagem que escolheu Média geométrica: " + str(np.std(porcentagem_escolheu_G_MEAN)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_G_MEAN[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_G_MEAN)
            lin_inserir = lin_inserir + 1
            #print("")
            '''
            print("Média de porcentagem que escolheu Máximo: " + str(np.mean(porcentagem_escolheu_MAX)))
            print("Média de porcentagem que escolheu Mínimo: " + str(np.mean(porcentagem_escolheu_MIN)))
            print("Média de porcentagem que escolheu Votação ponderada no comitê: " + str(np.mean(porcentagem_escolheu_sum_weight)))
            print("Média de porcentagem que escolheu Peso ponderado no comitê: " + str(np.mean(porcentagem_escolheu_peso_ponderado_comite)))
            '''


            print("Média de porcentagem que escolheu Rede Neural HARD: " + str(np.mean(porcentagem_escolheu_rede_neural[ii])))
            #print("Desvio padrão de porcentagem que escolheu Rede Neural HARD: " + str(np.std(porcentagem_escolheu_rede_neural)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_rede_neural[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_rede_neural)
            lin_inserir = lin_inserir + 1
            #print("")

            print("Média de porcentagem que escolheu Rede Neural SOFT: " + str(np.mean(porcentagem_escolheu_rede_neural_soft[ii])))
            #print("Desvio padrão de porcentagem que escolheu Rede Neural SOFT: " + str(np.std(porcentagem_escolheu_rede_neural_soft)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_rede_neural_soft[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_rede_neural_soft)
            lin_inserir = lin_inserir + 1
            #print("")

            
            #print("Média de porcentagem que escolheu Rede Neural SOFT class: " + str(np.mean(porcentagem_escolheu_rede_neural_soft_div)))
            #print("Desvio padrão de porcentagem que escolheu Rede Neural SOFT class: " + str(np.std(porcentagem_escolheu_rede_neural_soft_div)))
            #ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_rede_neural_soft_div)
            #ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_rede_neural_soft_div)
            lin_inserir = lin_inserir + 1
            #print("")
            

            print("Média de porcentagem que escolheu borda: " + str(np.mean(porcentagem_escolheu_borda[ii])))
            #print("Desvio padrão de porcentagem que escolheu borda: " + str(np.std(porcentagem_escolheu_borda)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_borda[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_borda)
            lin_inserir = lin_inserir + 1
            #print("")

            print("Média de porcentagem que escolheu naive bayes: " + str(np.mean(porcentagem_escolheu_naive_bayes[ii])))
            #print("Desvio padrão de porcentagem que escolheu naive bayes: " + str(np.std(porcentagem_escolheu_naive_bayes)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolheu_naive_bayes[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_naive_bayes)
            lin_inserir = lin_inserir + 1

            print("Média de porcentagem que escolheu comitês conjuntos: " + str(np.mean(porcentagem_escolha_compartilhada[ii])))
            #print("Desvio padrão de porcentagem que escolheu naive bayes: " + str(np.std(porcentagem_escolheu_naive_bayes)))
            ws[col_valor + str(lin_inserir)] = np.mean(porcentagem_escolha_compartilhada[ii])
            ws[col_desvio + str(lin_inserir)] = "-" #np.std(porcentagem_escolheu_naive_bayes)
            lin_inserir = lin_inserir + 1  
            
           
            print("")
            print("")
            print("Resultados maximo na combinacao")
            #print(resultados_hard)
            print("Média: " + str(np.mean(resultados_maximo_na_combinacao)*100))
            print("Desvio padrão: " + str(np.std(resultados_maximo_na_combinacao)*100))
            print("")
            ws[col_valor + str(lin_inserir)] = np.mean(resultados_maximo_na_combinacao)*100
            ws[col_desvio + str(lin_inserir)] = np.std(resultados_maximo_na_combinacao)*100
            lin_inserir = lin_inserir + 1

            print("Média de amostras de val_1 com empate de classificações inicio da região de competência: " + str(np.mean(geral_empate_classificadores_no_inicio[ii])))
            print("Desvio padrão: " + str(np.std(geral_empate_classificadores_no_inicio[ii])))
            ws[col_valor + str(lin_inserir)] = np.mean(geral_empate_classificadores_no_inicio[ii])
            ws[col_desvio + str(lin_inserir)] = np.std(geral_empate_classificadores_no_inicio[ii])

            lin_inserir = lin_inserir + 1
            print("")
            print("")
            
            print("Média de amostras de val_1 com empate de classificações em toda a região de competência: " + str(np.mean(geral_empate_classificadores_no_final[ii])))
            print("Desvio padrão : " + str(np.std(geral_empate_classificadores_no_final[ii])))
            ws[col_valor + str(lin_inserir)] = np.mean(geral_empate_classificadores_no_final[ii])
            ws[col_desvio + str(lin_inserir)] = np.std(geral_empate_classificadores_no_final[ii])
            lin_inserir = lin_inserir + 1

            print("")
            print("")

            lin_inserir = lin_inserir - 12
            coluna_valor = coluna_valor + 2
            coluna_desvio_padrao = coluna_desvio_padrao +2
            col_valor = get_column_letter(coluna_valor)
            col_desvio = get_column_letter(coluna_desvio_padrao)
            
        #coluna_valor = coluna_valor + 2
        #coluna_desvio_padrao = coluna_desvio_padrao + 2


        print("^^^^--------------------------- FIM  K = "+ str(melhor_k)+" ---------------------------^^")


wb.save('comite estatico NOVA FUNCAO.xlsx')
print("ARQUIVO SALVO")
